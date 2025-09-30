import streamlit as st
import pandas as pd
import io
import numpy as np
import locale
from datetime import datetime
import requests
from msal import ConfidentialClientApplication
import zipfile
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows

# ==============================================================================
# CONFIGURACI√ìN DE SHAREPOINT Y AZURE
# ==============================================================================
CLIENT_ID = "b469ba00-b7b6-434c-91bf-d3481c171da5"
CLIENT_SECRET = "8nS8Q~tAYqkeISRUQyOBBAsLn6b_Z8LdNQR23dnn"
TENANT_ID = "f20cbde7-1c45-44a0-89c5-63a25c557ef8"
SHAREPOINT_HOSTNAME = "iacsas.sharepoint.com"
SITE_NAME = "PruebasProyectosSantiago"
RUTA_CARPETA_VENTAS_MENSUALES = "Ventas con ciudad 2025"
# ==============================================================================
# FUNCIONES DE AUTENTICACI√ìN Y CONEXI√ìN
# ==============================================================================
AUTHORITY = f"https://login.microsoftonline.com/{TENANT_ID}"
SCOPES = ["https://graph.microsoft.com/.default"]

def actualizar_archivo_trm(headers, site_id, ruta_archivo_trm, df_datos_procesados, status_placeholder):
    """
    Actualiza la hoja "Datos" del TRM.xlsx a√±adiendo nuevas filas sin borrar las existentes,
    lo que permite a las Tablas de Excel extender las f√≥rmulas autom√°ticamente.
    """
    nombre_hoja_destino = "Datos"
    status_placeholder.info(f"üîÑ Iniciando actualizaci√≥n (modo ap√©ndice) de la hoja '{nombre_hoja_destino}'...")

    try:
        # PASO 1: Descargar y cargar el libro de trabajo completo
        status_placeholder.info("1/3 - Descargando archivo TRM...")
        contenido_trm_bytes = obtener_contenido_archivo_sharepoint(headers, site_id, ruta_archivo_trm)
        if contenido_trm_bytes is None: return False

        libro = openpyxl.load_workbook(io.BytesIO(contenido_trm_bytes))
        if nombre_hoja_destino not in libro.sheetnames:
            status_placeholder.error(f"‚ùå No se encontr√≥ la hoja '{nombre_hoja_destino}'.")
            return False
        
        hoja = libro[nombre_hoja_destino]

        # PASO 2: Preparar las nuevas filas para ser a√±adidas
        status_placeholder.info("2/3 - Preparando nuevas filas para a√±adir...")
        
        # Obtener el n√∫mero de encabezados de la hoja de destino
        num_encabezados = len([cell.value for cell in hoja[1]])
        
        lista_nuevas_filas = []
        # Iterar sobre las filas de los datos procesados
        for index, fila_procesada in df_datos_procesados.iterrows():
            # Crear una lista de strings vac√≠os del tama√±o de la fila de destino
            nueva_fila_lista = [""] * num_encabezados
            
            # --- CORRECCI√ìN Y SIMPLIFICACI√ìN AQU√ç ---
            # Copiar los valores de la fila procesada a la nueva lista, a partir de la 4ta posici√≥n (√≠ndice 3)
            for i, valor in enumerate(fila_procesada.values):
                if (i + 3) < num_encabezados:
                    nueva_fila_lista[i + 3] = valor
            
            lista_nuevas_filas.append(nueva_fila_lista)

        # PASO 3: A√±adir las nuevas filas y subir el archivo
        status_placeholder.info(f"3/3 - A√±adiendo {len(lista_nuevas_filas)} nuevas filas y subiendo...")
        for fila in lista_nuevas_filas:
            hoja.append(fila) # 'append' a√±ade la fila al final, sin tocar las existentes
            
        # Guardar y subir
        output = io.BytesIO()
        libro.save(output)
        
        endpoint_put = f"https://graph.microsoft.com/v1.0/sites/{site_id}/drive/root:/{ruta_archivo_trm}:/content"
        response_put = requests.put(endpoint_put, data=output.getvalue(), headers=headers)
        response_put.raise_for_status()

        status_placeholder.success("‚úÖ ¬°Archivo TRM actualizado! Las f√≥rmulas deber√≠an haberse extendido autom√°ticamente.")
        return True

    except Exception as e:
        status_placeholder.error(f"‚ùå Fall√≥ la actualizaci√≥n del archivo TRM. Error: {e}")
        return False

'''
def actualizar_archivo_trm(headers, site_id, ruta_archivo_trm, df_datos_procesados, status_placeholder):
    """
    Actualiza la hoja "Datos" del TRM.xlsx de forma segura, preservando formatos y hojas,
    y utilizando las funciones de validaci√≥n para evitar errores con archivos corruptos.
    """
    nombre_hoja_destino = "Datos"
    status_placeholder.info(f"üîÑ Iniciando actualizaci√≥n avanzada de la hoja '{nombre_hoja_destino}'...")

    try:
        # PASO 1: Verificar que el archivo TRM existe y es accesible (REUTILIZANDO TUS FUNCIONES)
        status_placeholder.info("1/5 - Verificando metadatos del archivo TRM...")
        existe, metadata = verificar_archivo_existe_sharepoint(headers, site_id, ruta_archivo_trm)
        if not existe:
            status_placeholder.error("‚ùå No se puede continuar: el archivo TRM no fue encontrado.")
            return False

        # PASO 2: Descargar el contenido con validaciones avanzadas (REUTILIZANDO TUS FUNCIONES)
        status_placeholder.info("2/5 - Descargando y validando el contenido del archivo TRM...")
        contenido_trm_bytes = obtener_contenido_archivo_sharepoint(headers, site_id, ruta_archivo_trm)
        if contenido_trm_bytes is None:
            status_placeholder.error("‚ùå Falla en la descarga o validaci√≥n del archivo TRM.")
            return False
        
        contenido_trm = io.BytesIO(contenido_trm_bytes)

        # PASO 3: Cargar el libro de trabajo con openpyxl (ahora sabemos que es seguro hacerlo)
        status_placeholder.info("3/5 - Modificando el archivo Excel en memoria...")
        libro = openpyxl.load_workbook(contenido_trm)
        
        if nombre_hoja_destino not in libro.sheetnames:
            status_placeholder.error(f"‚ùå No se encontr√≥ la hoja de destino '{nombre_hoja_destino}'.")
            return False
        
        # Leer los datos de la hoja espec√≠fica en un DataFrame
        df_trm_existente = pd.read_excel(io.BytesIO(contenido_trm_bytes), sheet_name=nombre_hoja_destino, engine='openpyxl')
        df_trm_existente.reset_index(drop=True, inplace=True)
        
        # PASO 4: Preparar los nuevos datos y combinarlos
        status_placeholder.info("4/5 - Preparando y combinando los nuevos datos...")
        # (Aqu√≠ va la misma l√≥gica que ya ten√≠amos para crear df_para_agregar y combinar)
        lista_nuevas_filas = []
        nombres_columnas_destino = df_trm_existente.columns
        nombres_columnas_origen = df_datos_procesados.columns
        for index, fila_origen in df_datos_procesados.iterrows():
            nueva_fila = {}
            nueva_fila[nombres_columnas_destino[0]] = ""
            nueva_fila[nombres_columnas_destino[1]] = ""
            nueva_fila[nombres_columnas_destino[2]] = ""
            for i, col_destino in enumerate(nombres_columnas_destino[3:]):
                if i < len(nombres_columnas_origen):
                    nueva_fila[col_destino] = fila_origen[nombres_columnas_origen[i]]
            lista_nuevas_filas.append(nueva_fila)
        
        df_para_agregar = pd.DataFrame(lista_nuevas_filas)
        df_trm_actualizado = pd.concat([df_trm_existente, df_para_agregar], ignore_index=True)
        
        # --- L√çNEAS NUEVAS PARA ELIMINAR DUPLICADOS ---
        filas_antes = len(df_trm_actualizado)
        df_sin_duplicados = df_trm_actualizado.drop_duplicates(keep='first')
        filas_despues = len(df_sin_duplicados)
        
        duplicados_encontrados = filas_antes - filas_despues
        if duplicados_encontrados > 0:
            status_placeholder.warning(f"‚ö†Ô∏è Se encontraron y omitieron {duplicados_encontrados} registros duplicados en el archivo TRM.")
        else:
            status_placeholder.info("‚úÖ No se encontraron registros duplicados en el archivo TRM.")
        
        # --- FIN DE L√çNEAS NUEVAS ---

        # Limpiar columnas "Unnamed:"
        cols_a_eliminar = [col for col in df_sin_duplicados.columns if 'Unnamed:' in str(col)]
        if cols_a_eliminar:
            df_sin_duplicados.drop(columns=cols_a_eliminar, inplace=True)

        # Escribir los datos actualizados en la hoja de openpyxl
        hoja = libro[nombre_hoja_destino]
        for r in range(hoja.max_row, 1, -1):
            hoja.delete_rows(r)
        from openpyxl.utils.dataframe import dataframe_to_rows
        for r_idx, row in enumerate(dataframe_to_rows(df_sin_duplicados, index=False, header=False), 2):
            for c_idx, value in enumerate(row, 1):
                hoja.cell(row=r_idx, column=c_idx, value=value)
                
        # PASO 5: Guardar y subir el libro modificado
        status_placeholder.info("5/5 - Subiendo el archivo final a SharePoint...")
        output = io.BytesIO()
        libro.save(output)
        
        endpoint_put = f"https://graph.microsoft.com/v1.0/sites/{site_id}/drive/root:/{ruta_archivo_trm}:/content"
        response_put = requests.put(endpoint_put, data=output.getvalue(), headers=headers)
        response_put.raise_for_status()

        #status_placeholder.success("‚úÖ ¬°Archivo TRM actualizado preservando f√≥rmulas y formatos!")
        return True

    except Exception as e:
        #status_placeholder.error(f"‚ùå Fall√≥ la actualizaci√≥n del archivo TRM. Error: {e}")
        return False
'''

def validar_respuesta_sharepoint(response, nombre_archivo):
    """
    Valida que la respuesta de SharePoint sea correcta y contenga un archivo Excel
    """
    #st.info(f"üîç Validando respuesta para: {nombre_archivo}")
    
    # 1. Verificar c√≥digo de estado HTTP
    #st.write(f"üìä C√≥digo HTTP: {response.status_code}")
    
    if response.status_code != 200:
        #st.error(f"‚ùå Error HTTP {response.status_code}")
        try:
            error_json = response.json()
            #st.json(error_json)
        except:
            st.error(f"Texto de respuesta: {response.text[:500]}...")
        return False, "Error HTTP"
    
    # 2. Verificar el tama√±o del contenido
    content_length = len(response.content)
    #st.write(f"üìè Tama√±o del archivo descargado: {content_length:,} bytes")
    
    if content_length == 0:
        st.error("‚ùå El archivo est√° vac√≠o (0 bytes)")
        return False, "Archivo vac√≠o"
    
    if content_length < 100:  # Un Excel v√°lido debe tener al menos algunos cientos de bytes
        #st.warning("‚ö†Ô∏è El archivo es muy peque√±o para ser un Excel v√°lido")
        #st.write(f"Contenido recibido: {response.content}")
        return False, "Archivo muy peque√±o"
    
    # 3. Verificar el Content-Type si est√° disponible
    content_type = response.headers.get('Content-Type', 'No especificado')
    #st.write(f"üìã Content-Type: {content_type}")
    
    # 4. Verificar las primeras bytes para asegurar que es un archivo Excel
    primeros_bytes = response.content[:20]
    #st.write(f"üî¢ Primeros 20 bytes (hex): {primeros_bytes.hex()}")
    
    # Un archivo Excel (.xlsx) debe comenzar con la signature de ZIP: "PK"
    if not response.content.startswith(b'PK'):
        #st.error("‚ùå El archivo no tiene la signature de un archivo ZIP/Excel v√°lido")
        #st.error("Los archivos .xlsx deben comenzar con 'PK' (signature de ZIP)")
        
        # Mostrar el inicio del contenido como texto para debug
        try:
            inicio_texto = response.content[:200].decode('utf-8', errors='ignore')
            #st.error(f"Inicio del contenido como texto: {inicio_texto}")
        except:
            st.error("No se pudo decodificar el inicio del contenido como texto")
        
        return False, "Signature inv√°lida"
    
    #st.success("‚úÖ El archivo parece ser un Excel v√°lido")
    return True, "V√°lido"

def obtener_contenido_archivo_sharepoint(headers, site_id, ruta_archivo):
    """
    Descarga un archivo espec√≠fico de SharePoint con validaciones completas
    """
    #st.info(f"üì• Descargando archivo: {ruta_archivo}")
    
    # Construir el endpoint
    endpoint_get = f"https://graph.microsoft.com/v1.0/sites/{site_id}/drive/root:/{ruta_archivo}:/content"
    #st.write(f"üîó Endpoint: {endpoint_get}")
    
    try:
        # Realizar la petici√≥n
        response_get = requests.get(endpoint_get, headers=headers)
        
        # Validar la respuesta
        es_valido, mensaje = validar_respuesta_sharepoint(response_get, ruta_archivo.split('/')[-1])
        
        if not es_valido:
            #st.error(f"‚ùå Validaci√≥n fall√≥: {mensaje}")
            return None
        
        return response_get.content
        
    except requests.exceptions.RequestException as e:
        #st.error(f"‚ùå Error de red al descargar el archivo: {e}")
        return None
    except Exception as e:
        #st.error(f"‚ùå Error inesperado: {e}")
        return None

def verificar_archivo_existe_sharepoint(headers, site_id, ruta_archivo):
    """
    Verifica si un archivo existe y obtiene sus metadatos antes de descargarlo
    """
    #st.info(f"üîç Verificando existencia de: {ruta_archivo}")
    
    # Endpoint para obtener metadatos del archivo (sin descargar el contenido)
    endpoint_metadata = f"https://graph.microsoft.com/v1.0/sites/{site_id}/drive/root:/{ruta_archivo}"
    
    try:
        response = requests.get(endpoint_metadata, headers=headers)
        
        if response.status_code == 200:
            metadata = response.json()
            
            nombre = metadata.get('name', 'Sin nombre')
            tamano = metadata.get('size', 0)
            tipo = metadata.get('file', {}).get('mimeType', 'No especificado')
            modificado = metadata.get('lastModifiedDateTime', 'No especificado')
            
            #st.success(f"‚úÖ Archivo encontrado: {nombre}")
            #st.write(f"üìè Tama√±o: {tamano:,} bytes")
            #st.write(f"üìã Tipo MIME: {tipo}")
            #st.write(f"üìÖ √öltima modificaci√≥n: {modificado}")
            
            # Verificar que sea realmente un archivo Excel
            if tipo and 'spreadsheet' not in tipo.lower() and 'excel' not in tipo.lower():
                st.warning(f"‚ö†Ô∏è Advertencia: El tipo MIME '{tipo}' no parece ser un Excel")
            
            return True, metadata
        else:
            #st.error(f"‚ùå Archivo no encontrado. HTTP {response.status_code}")
            try:
                error_json = response.json()
                #st.json(error_json)
            except:
                st.error(f"Respuesta: {response.text}")
            return False, None
            
    except Exception as e:
        #t.error(f"‚ùå Error al verificar archivo: {e}")
        return False, None


def get_access_token(status_placeholder):
    #status_placeholder.info("‚öôÔ∏è Paso 2/5: Autenticando con Microsoft...")
    app = ConfidentialClientApplication(
        client_id=CLIENT_ID,
        authority=AUTHORITY,
        client_credential=CLIENT_SECRET
    )
    result = app.acquire_token_for_client(scopes=SCOPES)
    if "access_token" in result:
        #st.success("‚úÖ Token de acceso obtenido con √©xito.")
        return result['access_token']
    else:
        #st.error(f"Error al obtener token: {result.get('error_description')}")
        return None

def get_sharepoint_site_id(access_token):
    headers = {'Authorization': f'Bearer {access_token}'}
    site_url = f"https://graph.microsoft.com/v1.0/sites/{SHAREPOINT_HOSTNAME}:/sites/{SITE_NAME}"
    try:
        response = requests.get(site_url, headers=headers)
        response.raise_for_status()
        site_id = response.json().get('id')
        #st.success(f"‚úÖ Conexi√≥n exitosa con el sitio SharePoint: '{SITE_NAME}'")
        return site_id
    except requests.exceptions.RequestException as e:
        #st.error(f"Error al obtener site_id: {e.response.text}")
        return None

def encontrar_archivo_del_mes(headers, site_id, ruta_carpeta, status_placeholder):
    """
    Busca dentro de una CARPETA espec√≠fica y devuelve la RUTA COMPLETA del archivo del mes.
    """
    try:
        # Meses en espa√±ol con diferentes variaciones
        fecha_actual = datetime.now()
        mes_numero = fecha_actual.month
        
        # Diferentes patrones que podr√≠a tener el archivo
        patrones_busqueda = [
            f"{mes_numero}. ",  # "9. " para septiembre
            "Septiembre",       # Nombre completo del mes
            "septiembre",       # Min√∫scula
            f"{mes_numero:02d}",# "09" con cero delante
        ]
        
        #st.info(f"üîç Buscando archivo del mes {mes_numero} (Septiembre) en: '{ruta_carpeta}'")
        #st.write(f"Patrones de b√∫squeda: {patrones_busqueda}")
        
        # Primero, listar TODOS los archivos en la carpeta
        #st.write("üìÇ Listando todos los archivos disponibles:")
        endpoint_children = f"https://graph.microsoft.com/v1.0/sites/{site_id}/drive/root:/{ruta_carpeta}:/children"
        response_list = requests.get(endpoint_children, headers=headers)
        
        if response_list.status_code == 200:
            todos_archivos = response_list.json().get('value', [])
            
            #st.write(f"üìä Total de archivos en la carpeta: {len(todos_archivos)}")
            
            # Mostrar todos los archivos para debug
            for item in todos_archivos:
                if not item.get('folder'):  # Solo archivos, no carpetas
                    nombre = item.get('name', '')
                    tama√±o = item.get('size', 0)
                    #st.write(f"üìÑ {nombre} ({tama√±o:,} bytes)")
            
            # Buscar el archivo que coincida con los patrones
            archivos_candidatos = []
            
            for item in todos_archivos:
                if item.get('folder'):  # Saltar carpetas
                    continue
                    
                nombre_archivo = item.get('name', '').lower()
                
                # Verificar cada patr√≥n
                for patron in patrones_busqueda:
                    if patron.lower() in nombre_archivo:
                        archivos_candidatos.append({
                            'nombre_original': item.get('name'),
                            'ruta_completa': f"{ruta_carpeta}/{item.get('name')}",
                            'tama√±o': item.get('size', 0),
                            'patron_encontrado': patron
                        })
                        break  # Salir del loop de patrones una vez encontrado
            
            if archivos_candidatos:
                #st.success(f"‚úÖ Encontrados {len(archivos_candidatos)} archivos candidatos:")
                
                #for i, candidato in enumerate(archivos_candidatos):
                    #st.write(f"{i+1}. **{candidato['nombre_original']}** ({candidato['tama√±o']:,} bytes) - Patr√≥n: '{candidato['patron_encontrado']}'")
                
                # Seleccionar el primer candidato (o puedes agregar l√≥gica m√°s sofisticada)
                archivo_seleccionado = archivos_candidatos[0]
                #st.success(f"üéØ Archivo seleccionado: **{archivo_seleccionado['nombre_original']}**")
                
                return archivo_seleccionado['ruta_completa']
            else:
                #st.warning(f"‚ö†Ô∏è No se encontraron archivos que coincidan con los patrones para el mes {mes_numero}")
                
                # Mostrar sugerencia
                #st.info("üí° Archivos disponibles que podr√≠an ser relevantes:")
                for item in todos_archivos:
                    if not item.get('folder'):
                        nombre = item.get('name', '')
                        if any(char.isdigit() for char in nombre):  # Si contiene n√∫meros
                            st.write(f"ü§î {nombre}")
                
                return None
        else:
            st.error(f"‚ùå No se pudo listar el contenido de la carpeta. HTTP {response_list.status_code}")
            return None
            
    except requests.exceptions.RequestException as e:
        st.error(f"Error de conexi√≥n al buscar el archivo del mes: {e.response.text if e.response else e}")
        return None
    except Exception as e:
        st.error(f"Error inesperado durante la b√∫squeda del mes: {e}")
        return None

def agregar_datos_a_excel_sharepoint(headers, site_id, ruta_archivo, df_nuevos_datos, status_placeholder):
    """
    Agrega datos a la primera hoja de un archivo Excel en SharePoint,
    preservando f√≥rmulas, formatos y otras hojas.
    """
    #status_placeholder.info(f"üîÑ Iniciando actualizaci√≥n avanzada de: '{ruta_archivo.split('/')[-1]}'")

    try:
        # PASO 1: Descargar el archivo existente con validaciones
        #status_placeholder.info("1/4 - Descargando y validando archivo...")
        contenido_bytes = obtener_contenido_archivo_sharepoint(headers, site_id, ruta_archivo)
        if contenido_bytes is None:
            #status_placeholder.error("‚ùå Falla en la descarga o validaci√≥n del archivo.")
            return False

        contenido_en_memoria = io.BytesIO(contenido_bytes)

        # PASO 2: Cargar el libro de trabajo completo con openpyxl
        #status_placeholder.info("2/4 - Cargando estructura del archivo (formatos, hojas)...")
        libro = openpyxl.load_workbook(contenido_en_memoria)
        
        # Asumimos que los datos se agregan a la primera hoja.
        # Puedes cambiar esto por un nombre fijo si es necesario, ej: nombre_hoja_destino = "Ventas"
        nombre_hoja_destino = libro.sheetnames[0]
        hoja = libro[nombre_hoja_destino]
        
        # Leer los datos de esa hoja en un DataFrame para facilitar la manipulaci√≥n
        df_existente = pd.read_excel(io.BytesIO(contenido_bytes), sheet_name=nombre_hoja_destino, engine='openpyxl')
        df_existente.reset_index(drop=True, inplace=True)

        # PASO 3: Combinar los datos y limpiar columnas "Unnamed"
        #status_placeholder.info("3/4 - Combinando datos nuevos y existentes...")
        df_combinado = pd.concat([df_existente, df_nuevos_datos], ignore_index=True)
        
        # --- L√çNEAS NUEVAS PARA ELIMINAR DUPLICADOS ---
        filas_antes = len(df_combinado)
        # Elimina filas que son completamente id√©nticas, manteniendo la primera aparici√≥n
        df_sin_duplicados = df_combinado.drop_duplicates(keep='first')
        filas_despues = len(df_sin_duplicados)
        
        duplicados_encontrados = filas_antes - filas_despues
        if duplicados_encontrados > 0:
            status_placeholder.warning(f"‚ö†Ô∏è Se encontraron y omitieron {duplicados_encontrados} registros duplicados.")
        else:
            status_placeholder.info("‚úÖ No se encontraron registros duplicados.")
        
        # --- FIN DE L√çNEAS NUEVAS ---
        
        cols_a_eliminar = [col for col in df_sin_duplicados.columns if 'Unnamed:' in str(col)]
        if cols_a_eliminar:
            df_sin_duplicados.drop(columns=cols_a_eliminar, inplace=True)
            #status_placeholder.info("üßπ Columnas 'Unnamed:' eliminadas.")

        # PASO 4: Escribir los datos actualizados de vuelta a la hoja, preservando el resto
        #status_placeholder.info("4/4 - Escribiendo datos y subiendo el archivo final...")
        
        # Aseg√∫rate de usar el DataFrame limpio para el resto del proceso
        df_combinado = df_sin_duplicados # <-- ¬°Importante!
        hoja = libro[nombre_hoja_destino]
        # Borrar datos antiguos de la hoja (excepto encabezados) para evitar duplicados
        for r in range(hoja.max_row, 1, -1):
            hoja.delete_rows(r)
            
        from openpyxl.utils.dataframe import dataframe_to_rows    
        # Escribir el contenido del DataFrame combinado en la hoja
        for r_idx, row in enumerate(dataframe_to_rows(df_sin_duplicados, index=False, header=False), 2):
            for c_idx, value in enumerate(row, 1):
                hoja.cell(row=r_idx, column=c_idx, value=value)
        
        # Guardar el libro modificado en memoria
        output = io.BytesIO()
        libro.save(output)
        
        # Subir el archivo final
        endpoint_put = f"https://graph.microsoft.com/v1.0/sites/{site_id}/drive/root:/{ruta_archivo}:/content"
        response_put = requests.put(endpoint_put, data=output.getvalue(), headers=headers)
        response_put.raise_for_status()

        #status_placeholder.success(f"‚úÖ ¬°Archivo '{ruta_archivo.split('/')[-1]}' actualizado preservando su formato!")
        return True

    except Exception as e:
        status_placeholder.error(f"‚ùå Fall√≥ la actualizaci√≥n del archivo. Error: {e}")
        return False

def listar_archivos_en_carpeta(headers, site_id, ruta_carpeta):
    """
    Lista todos los archivos en una carpeta para debug
    """
    #st.info(f"üìÇ Explorando carpeta: {ruta_carpeta}")
    
    endpoint = f"https://graph.microsoft.com/v1.0/sites/{site_id}/drive/root:/{ruta_carpeta}:/children"
    
    try:
        response = requests.get(endpoint, headers=headers)
        if response.status_code == 200:
            items = response.json().get('value', [])
            
            #st.write(f"üìä Encontrados {len(items)} elementos:")
            for item in items:
                tipo = "üìÅ" if item.get('folder') else "üìÑ"
                nombre = item.get('name', 'Sin nombre')
                tamano = item.get('size', 0)
                #st.write(f"{tipo} {nombre} ({tamano:,} bytes)")
        else:
            st.error(f"‚ùå No se pudo listar la carpeta. HTTP {response.status_code}")
    except Exception as e:
        st.error(f"‚ùå Error: {e}")
    
    
# --- Funci√≥n Principal de Procesamiento ---
def procesar_excel_para_streamlit(uploaded_file, status_placeholder):
    """
    Procesa el archivo de Excel subido:
    - Ignora las primeras 7 filas al cargar el archivo (asumiendo que los encabezados est√°n en la fila 8).
    - Elimina filas con 'Tipo clasificaci√≥n' vac√≠o.
    - Elimina columnas no deseadas.
    - Actualiza la columna 'Total'.
    - Rellena 'Tasa de cambio' con TRM de API bajo condiciones espec√≠ficas.

    Args:
        uploaded_file (streamlit.UploadedFile): El archivo Excel subido por el usuario.

    Returns:
        pandas.DataFrame or None: El DataFrame procesado o None si hay un error.
    """
    try:
        # Usar skiprows para que Pandas lea el encabezado correcto
        df = pd.read_excel(uploaded_file, skiprows=7) # La fila 8 (√≠ndice 7) se toma como encabezado

        # Verifica si el DataFrame tiene columnas despu√©s de skiprows.
        if df.empty or df.columns.empty:
            st.error("Parece que el archivo no tiene datos o encabezados despu√©s de saltar las primeras 7 filas. Por favor, verifica el formato del archivo.")
            return None

        st.info(f"Archivo cargado exitosamente. Se saltaron las primeras 7 filas. Filas iniciales (despu√©s de saltar): **{len(df)}**.")

        df_procesado = df.copy()
        
        # --- FUNCI√ìN DE LIMPIEZA SIMPLE ---
        def convertir_a_numero_limpiando_comas(columna):
            if not pd.api.types.is_string_dtype(columna):
                columna = columna.astype(str)
            columna_limpia = columna.str.replace(',', '', regex=False)
            return pd.to_numeric(columna_limpia, errors='coerce')

        if 'Identificaci√≥n Vendedor' in df_procesado.columns:
            # Crear la nueva columna 'Vendedor' con los datos de la original
            df_procesado['Vendedor'] = df_procesado['Identificaci√≥n Vendedor']
            st.success("‚úÖ Columna 'Vendedor' creada con √©xito.")
        else:
            # Si la columna original no existe, crear 'Vendedor' como una columna vac√≠a
            st.warning("‚ö†Ô∏è No se encontr√≥ la columna 'Identificaci√≥n Vendedor'. Se crear√° una columna 'Vendedor' vac√≠a.")
            df_procesado['Vendedor'] = ''

        # Columnas a eliminar predefinidas
        nombres_columnas_a_eliminar = [
            "Sucursal",
            "Centro costo",
            "Fecha creaci√≥n",
            "Fecha modificaci√≥n",
            "Correo electr√≥nico",
            "Tipo de registro",
            "Referencia f√°brica",
            "Bodega",
            #"Identificaci√≥n Vendedor",
            "Nombre vendedor",
            "Valor desc.",
            "Base AIU",
            "Impuesto cargo",
            "Valor Impuesto Cargo",
            "Impuesto Cargo 2",
            "Valor Impuesto Cargo 2",
            "Impuesto retenci√≥n",
            "Valor Impuesto Retenci√≥n",
            "Base retenci√≥n (ICA/IVA)",
            "Cargo en totales",
            "Descuento en totales",
            "Moneda",
            "Forma pago",
            "Fecha vencimiento",
            "Nombre contacto"
        ]

        #df_procesado = df.copy()

        # 1. Eliminar filas donde "Tipo clasificaci√≥n" est√© vac√≠o/NaN
        if "Tipo clasificaci√≥n" in df_procesado.columns:
            filas_antes_eliminacion = len(df_procesado)
            df_procesado.dropna(subset=["Tipo clasificaci√≥n"], inplace=True)
            filas_despues_eliminacion = len(df_procesado)
            st.success(f"Filas con 'Tipo clasificaci√≥n' vac√≠o eliminadas: **{filas_antes_eliminacion - filas_despues_eliminacion}**. Filas restantes: **{filas_despues_eliminacion}**.")
        else:
            st.warning("La columna **'Tipo clasificaci√≥n'** no se encontr√≥. No se eliminaron filas vac√≠as.")

        # 2. Eliminar columnas especificadas
        columnas_existentes_para_eliminar = [col for col in nombres_columnas_a_eliminar if col in df_procesado.columns]
        columnas_no_existentes_para_eliminar = [col for col in nombres_columnas_a_eliminar if col not in df_procesado.columns]

        if columnas_existentes_para_eliminar:
            df_procesado.drop(columns=columnas_existentes_para_eliminar, inplace=True)
            st.success(f"Columnas eliminadas: **{', '.join(columnas_existentes_para_eliminar)}**.")
        else:
            st.info("Ninguna de las columnas especificadas para eliminar se encontr√≥. No se eliminaron columnas.")

        if columnas_no_existentes_para_eliminar:
            st.warning(f"Advertencia: Las siguientes columnas especificadas para eliminaci√≥n no se encontraron: **{', '.join(columnas_no_existentes_para_eliminar)}**.")

        # 3. Actualizar la columna "Total" existente
        if "Cantidad" in df_procesado.columns and "Valor unitario" in df_procesado.columns and "Total" in df_procesado.columns:
            df_procesado["Cantidad"] = pd.to_numeric(df_procesado["Cantidad"], errors='coerce')
            df_procesado["Valor unitario"] = pd.to_numeric(df_procesado["Valor unitario"], errors='coerce')
            df_procesado["Total"] = df_procesado["Cantidad"] * df_procesado["Valor unitario"]
            df_procesado["Total"] = df_procesado["Total"].fillna(0)
            st.success("La columna **'Total'** ha sido actualizada con el c√°lculo **'Cantidad * Valor unitario'**.")
        else:
            st.warning("Advertencia: No se pudieron encontrar las columnas **'Cantidad'**, **'Valor unitario'** y/o **'Total'**. La columna **'Total'** no se actualiz√≥.")

        # 4. Crear y posicionar la nueva columna "Numero comprobante"
        columnas_necesarias = ['N√∫mero comprobante', 'Consecutivo', 'Factura proveedor']
        if all(col in df_procesado.columns for col in columnas_necesarias):
            # Definir las condiciones
            conditions = [
                df_procesado['N√∫mero comprobante'] == 'FV-1',
                df_procesado['N√∫mero comprobante'] == 'FV-2'
            ]
            
            # Definir los valores a asignar para cada condici√≥n
            choices = [
                'FLE-' + df_procesado['Consecutivo'].astype('Int64').astype(str),
                'FSE-' + df_procesado['Consecutivo'].astype('Int64').astype(str)
            ]
            
            # Usar np.select para crear los valores de la nueva columna
            # El valor por defecto ser√° un texto vac√≠o ''
            valores_nueva_columna = np.select(conditions, choices, default='')
            
            # Encontrar la posici√≥n de la columna "Factura proveedor" para insertar antes
            posicion_insercion = df_procesado.columns.get_loc('Factura proveedor')
            
            # Insertar la nueva columna en la posici√≥n encontrada
            df_procesado.insert(posicion_insercion, 'Numero comprobante', valores_nueva_columna)
            
            st.success("Se ha creado y llenado la nueva columna **'Numero comprobante'**.")
            
        else:
            st.warning("Advertencia: No se encontraron las columnas necesarias ('N√∫mero comprobante', 'Consecutivo', 'Factura proveedor') para crear la nueva columna.")
        
        # 5. Extraer TRM de 'Observaciones' y sobrescribir 'Tasa de cambio'
        #if "Tasa de cambio" in df_procesado.columns and "Observaciones" in df_procesado.columns:
            #st.info("Actualizando 'Tasa de cambio' con los valores encontrados en 'Observaciones'...")

            #df_procesado['Observaciones'] = df_procesado['Observaciones'].astype(str)
            # Extrae el contenido de las llaves '{}'. El resultado ser√° el texto o NaN si no hay llaves.
            #trm_extraida = df_procesado['Observaciones'].str.extract(r'\{(.*?)\}')[0]
            # Elimina las filas donde no se encontr√≥ nada (NaN), para quedarnos solo con los valores a actualizar.
            #trm_extraida.dropna(inplace=True)
            # Aseguramos que la columna 'Tasa de cambio' pueda recibir texto sin problemas.
            #df_procesado['Tasa de cambio'] = df_procesado['Tasa de cambio'].astype(object)
            # Actualiza la columna 'Tasa de cambio' S√ìLO con los valores encontrados.
            # El m√©todo .update() alinea por √≠ndice y solo modifica donde hay coincidencia.
            #df_procesado['Tasa de cambio'].update(trm_extraida)
            
            #filas_actualizadas = len(trm_extraida)
            #st.success(f"Se actualizaron **{filas_actualizadas}** filas en 'Tasa de cambio'. Los valores existentes se respetaron donde no se encontr√≥ un valor entre {{}}.")
        #else:
            #st.warning("Advertencia: No se encontraron las columnas **'Tasa de cambio'** y/o **'Observaciones'**.")
        # 5. Extraer, LIMPIAR y sobrescribir 'Tasa de cambio' desde 'Observaciones' (L√ìGICA CORREGIDA Y ENFOCADA)
        if "Tasa de cambio" in df_procesado.columns and "Observaciones" in df_procesado.columns:
            
            # Para evitar problemas, nos aseguramos de que la columna 'Tasa de cambio' sea num√©rica desde el principio.
            # Usamos la limpieza simple de comas que ya definimos.
            df_procesado['Tasa de cambio'] = convertir_a_numero_limpiando_comas(df_procesado['Tasa de cambio']).fillna(0)

            # 1. EXTRAER el valor de las observaciones como texto.
            trm_extraida = df_procesado['Observaciones'].astype(str).str.extract(r'\{(.*?)\}')[0]
            
            # Quitamos las filas donde no se encontr√≥ nada.
            trm_extraida.dropna(inplace=True)

            if not trm_extraida.empty:
                st.info("Valores de TRM encontrados en 'Observaciones'. Limpiando y actualizando...")

                # 2. LIMPIAR el texto extra√≠do (quitamos comas de miles).
                # Ejemplo: "4,061.36" se convierte en "4061.36"
                trm_limpia = trm_extraida.str.replace(',', '', regex=False)

                # 3. CONVERTIR el texto limpio a un formato num√©rico.
                trm_numerica = pd.to_numeric(trm_limpia, errors='coerce')
                
                # Quitamos las filas donde la conversi√≥n a n√∫mero pudo haber fallado.
                trm_numerica.dropna(inplace=True)

                # 4. ACTUALIZAR la columna 'Tasa de cambio' con los valores ya num√©ricos y limpios.
                # El m√©todo .update() alinea por √≠ndice y solo modifica donde encuentra correspondencia.
                df_procesado['Tasa de cambio'].update(trm_numerica)
                st.success(f"Se actualizaron **{len(trm_numerica)}** filas en 'Tasa de cambio' con valores num√©ricos limpios desde 'Observaciones'.")


        # 5.1. Calcular la nueva columna 'Valor Total ME' (VERSI√ìN CORREGIDA FINAL)
        st.info("Calculando 'Valor Total ME'...")
        if 'Total' in df_procesado.columns and 'Tasa de cambio' in df_procesado.columns:
            
            # PASO CLAVE: Nos aseguramos de que 'Tasa de cambio' sea num√©rica OTRA VEZ,
            # justo antes de la divisi√≥n, para revertir el cambio a 'object' del paso anterior.
            tasa_numerica = pd.to_numeric(df_procesado['Tasa de cambio'], errors='coerce')
            
            # Reemplazamos 0 con NaN para evitar errores de divisi√≥n por cero.
            tasa_numerica.replace(0, np.nan, inplace=True)

            # Realizamos la divisi√≥n.
            df_procesado['Valor Total ME'] = df_procesado['Total'] / tasa_numerica
            
            # Rellenamos cualquier resultado inv√°lido (NaN) con 0.
            df_procesado['Valor Total ME'].fillna(0, inplace=True)
            
            st.success("Se ha creado y calculado la columna **'Valor Total ME'**.")
        else:
            st.warning("No se pudo calcular 'Valor Total ME'.")

        # 6. Relacionar documentos FV-1 con DS-1 y FC-1
        st.info("Iniciando el proceso de relacionamiento de documentos...")
        
        # Separar el DataFrame en los dos grupos principales
        df_destino = df_procesado[df_procesado['N√∫mero comprobante'].isin(['FV-1', 'FV-2'])].copy()
        df_fuente = df_procesado[df_procesado['N√∫mero comprobante'].isin(['DS-1', 'FC-1'])].copy()

        if not df_fuente.empty:
            # Preparar el DataFrame fuente (DS-1, FC-1)
            df_fuente['NIT_relacion'] = df_fuente['Observaciones'].str.extract(r'\((.*?)\)')[0]
            
            df_destino['Identificaci√≥n'] = df_destino['Identificaci√≥n'].astype('Int64').astype(str)
            df_destino['C√≥digo'] = df_destino['C√≥digo'].astype(str)
            
            df_fuente['NIT_relacion'] = df_fuente['NIT_relacion'].astype(str)
            df_fuente['C√≥digo'] = df_fuente['C√≥digo'].astype(str)
            
            # A√±adir prefijo a las columnas para evitar colisiones y dar claridad
            df_fuente = df_fuente.add_prefix('REL_')
            
            # Realizar la uni√≥n externa (outer join)
            df_final = pd.merge(
                df_destino,
                df_fuente,
                how='outer',
                left_on=['Identificaci√≥n', 'C√≥digo'],
                right_on=['REL_NIT_relacion', 'REL_C√≥digo']
            )
            
            st.success("Relacionamiento completado. Los documentos sin pareja se han conservado.")
            df_procesado = df_final
        else:
            st.warning("No se encontraron documentos DS-1 o FC-1 para relacionar. El archivo final no tendr√° columnas de relaci√≥n.")
        
        # 7. Organizar y Limpiar Columnas Finales
        st.info("Organizando el formato final del archivo...")
        
        # A. Renombrar la columna "Tipo clasificaci√≥n" a "Tipo Bien"
        # Verificamos si la columna existe antes de intentar renombrarla
        if "Tipo clasificaci√≥n" in df_procesado.columns:
            df_procesado.rename(columns={"Tipo clasificaci√≥n": "Tipo Bien"}, inplace=True)
            st.info("La columna **'Tipo clasificaci√≥n'** ha sido renombrada a **'Tipo Bien'**.")
        
        if 'Tipo Bien' in df_procesado.columns:
            # Creamos un diccionario con los valores a reemplazar
            mapeo_valores = {
                'Servicio': 'S',
                'Producto': 'P'
            }
            df_procesado['Tipo Bien'].replace(mapeo_valores, inplace=True)
            st.info("Valores en 'Tipo Bien' actualizados: 'Servicio' a 'S' y 'Producto' a 'P'.")
        
        #Creaci√≥n de la nueva columna "Vendedor"
        #if 'Vendedor' not in df_procesado.columns:
            #df_procesado['Vendedor'] = ''
        if 'Identificaci√≥n Vendedor' in df_procesado.columns:
            df_procesado.drop(columns=['Identificaci√≥n Vendedor'], inplace=True)
            
        #Creaci√≥n de la nueva columna "Clasificaci√≥n Producto"
        if 'Clasificaci√≥n Producto' not in df_procesado.columns:
            df_procesado['Clasificaci√≥n Producto'] = ''
            
        #Creaci√≥n de la nueva columna "L√≠nea"
        if 'L√≠nea' not in df_procesado.columns:
            df_procesado['L√≠nea'] = ''
            
        #Creaci√≥n de la nueva columna "Descripci√≥n L√≠nea"
        if 'Descripci√≥n L√≠nea' not in df_procesado.columns:
            df_procesado['Descripci√≥n L√≠nea'] = ''
            
        #Creaci√≥n de la nueva columna "Subl√≠nea"
        if 'Subl√≠nea' not in df_procesado.columns:
            df_procesado['Subl√≠nea'] = ''
            
        #Creaci√≥n de la nueva columna "Descripci√≥n Subl√≠nea"
        if 'Descripci√≥n Subl√≠nea' not in df_procesado.columns:
            df_procesado['Descripci√≥n Subl√≠nea'] = ''
            
        
        #Se define el orden y la selecci√≥n final de las columnas
        columnas_finales = [
            # Columnas del lado izquierdo (FV)
            'Tipo Bien', 'Clasificaci√≥n Producto', 'L√≠nea', 'Descripci√≥n L√≠nea', 'Subl√≠nea', 'Descripci√≥n Subl√≠nea', 'C√≥digo', 'Nombre', 'N√∫mero comprobante', 'Numero comprobante',
            'Fecha elaboraci√≥n', 'Identificaci√≥n', 'Nombre tercero', 'Vendedor', 'Cantidad',
            'Valor unitario', 'Total', 'Tasa de cambio', 'Valor Total ME', 'Observaciones',
            
            # Columnas del lado derecho (REL_)
            'REL_N√∫mero comprobante', 'REL_Consecutivo',
            'REL_Factura proveedor', 'REL_Identificaci√≥n', 'REL_Nombre tercero', 'REL_Cantidad',
            'REL_Valor unitario',  'REL_Tasa de cambio', 'REL_Total', 'REL_Valor Total ME'
        ]
        
        # Filtrar la lista para incluir solo las columnas que realmente existen en el DataFrame
        # Esto hace el c√≥digo m√°s robusto si alguna columna faltara
        columnas_existentes_ordenadas = [col for col in columnas_finales if col in df_procesado.columns]

        # Reordenar y eliminar las columnas no deseadas de una sola vez
        df_procesado = df_procesado[columnas_existentes_ordenadas]

        st.success("Columnas reorganizadas y limpiadas con √©xito.")
 
        st.success("¬°Procesamiento completado con √©xito!")
        
        return df_procesado

    except Exception as e:
        st.error(f"Se produjo un error durante el procesamiento: {e}")
        return None

# --- Interfaz de Usuario de Streamlit ---
st.set_page_config(page_title="Procesador de Excel Autom√°tico", layout="centered")

st.title("üìä Procesador de Archivos Excel")
st.markdown("---")

uploaded_file = st.file_uploader(
    "Sube tu archivo Excel (.xlsx)",
    type=["xlsx"],
    help="Arrastra y suelta tu archivo Excel aqu√≠ o haz clic para buscar."
)

#st.markdown("---")
#st.header("üîß Herramientas de Debug para SharePoint")

# Crear variables de prueba para conexi√≥n SharePoint
#if st.button("üîó Probar Conexi√≥n SharePoint (Solo Debug)"):
    #with st.spinner("Conectando..."):
        #status_placeholder = st.empty()
        #token = get_access_token(status_placeholder)
        
        #if token:
            #site_id = get_sharepoint_site_id(token)
            #if site_id:
                #headers = {'Authorization': f'Bearer {token}'}
                #st.session_state.debug_headers = headers
                #st.session_state.debug_site_id = site_id
                #st.success("‚úÖ Conexi√≥n establecida para debug")

# Solo mostrar herramientas de debug si hay conexi√≥n
if hasattr(st.session_state, 'debug_headers') and hasattr(st.session_state, 'debug_site_id'):
    
    with st.expander("üß™ Debug de Archivos SharePoint", expanded=False):
        
        # Debug para la carpeta mensual
        #st.subheader("üìÖ Debug de Carpeta Mensual")
        if st.button("Listar archivos en carpeta mensual"):
            listar_archivos_en_carpeta(st.session_state.debug_headers, st.session_state.debug_site_id, "Ventas con ciudad 2025")
        
        # Debug para archivo espec√≠fico
        #st.subheader("üîç Debug de Archivo Espec√≠fico")
        archivo_debug = st.text_input("Ruta completa del archivo a verificar:", 
                                     "Ventas con ciudad 2025/Ventas Septiembre 2025.xlsx")
        
        col1, col2 = st.columns(2)
        with col1:
            if st.button("Verificar archivo"):
                if archivo_debug:
                    verificar_archivo_existe_sharepoint(st.session_state.debug_headers, st.session_state.debug_site_id, archivo_debug)
        
        with col2:
            if st.button("Intentar descargar"):
                if archivo_debug:
                    contenido = obtener_contenido_archivo_sharepoint(st.session_state.debug_headers, st.session_state.debug_site_id, archivo_debug)
                    if contenido:
                        st.success(f"Archivo descargado exitosamente ({len(contenido):,} bytes)")

st.markdown("---")

df_result = None


if uploaded_file is not None:
    st.success(f"Archivo **'{uploaded_file.name}'** cargado correctamente.")
    
    if st.button("Iniciar Procesamiento"):
        # Crear el placeholder una sola vez
        status_placeholder = st.empty()
        with st.spinner("Procesando tu archivo... Esto puede tardar unos minutos, especialmente al consultar la TRM..."):
            df_result = procesar_excel_para_streamlit(uploaded_file, status_placeholder)
        
        if df_result is not None:
            st.subheader("Vista previa del archivo procesado:")
            st.dataframe(df_result.head())

            output = io.BytesIO()
            # 2. Conectarse a SharePoint
            token = get_access_token(status_placeholder)
            if token:
                
                site_id = get_sharepoint_site_id(token) # Esta funci√≥n es r√°pida, no necesita placeholder

                if site_id:
                    # Una vez que tenemos el site_id, AHORA creamos los headers para las siguientes funciones
                    headers = {'Authorization': f'Bearer {token}'}
                    # 3. Encontrar el archivo del mes
                    ruta_archivo_mensual = encontrar_archivo_del_mes(headers, site_id, RUTA_CARPETA_VENTAS_MENSUALES, status_placeholder)
                    ruta_fija_trm = "01 Archivos Area Administrativa/TRM.xlsx"
                    exito_trm = actualizar_archivo_trm(headers, site_id, ruta_fija_trm, df_result, status_placeholder)
                    #st.info("Archivo TRM actualizado con √âxito")
                    if ruta_archivo_mensual:
                        # 4. Agregar los datos
                        #agregar_datos_a_excel_sharepoint(headers, site_id, ruta_archivo_mensual, df_result, status_placeholder)
                        exito = agregar_datos_a_excel_sharepoint(headers, site_id, ruta_archivo_mensual, df_result, status_placeholder)
                        if exito:
                            st.balloons()
            
            with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
                df_result.to_excel(writer, index=False, sheet_name='Procesado')
            processed_data = output.getvalue()

            st.download_button(
                label="Descargar Archivo Procesado",
                data=processed_data,
                file_name=f"procesado_{uploaded_file.name}",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            st.info("Tu archivo ha sido procesado y est√° listo para descargar.")
else:
    st.info("Por favor, sube un archivo Excel para comenzar.")
    
    



