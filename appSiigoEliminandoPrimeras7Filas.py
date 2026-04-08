import streamlit as st
import pandas as pd
import io
import numpy as np
import locale
from datetime import datetime
import requests
from msal import ConfidentialClientApplication
import zipfile
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import TableColumn
from openpyxl.utils import get_column_letter
from pandas.api.types import is_object_dtype

# ==============================================================================
# CONFIGURACIÓN DE SHAREPOINT Y AZURE
# ==============================================================================
CLIENT_ID = "b469ba00-b7b6-434c-91bf-d3481c171da5"
CLIENT_SECRET = "8nS8Q~tAYqkeISRUQyOBBAsLn6b_Z8LdNQR23dnn"
TENANT_ID = "f20cbde7-1c45-44a0-89c5-63a25c557ef8"
SHAREPOINT_HOSTNAME = "iacsas.sharepoint.com"
SITE_NAME = "PruebasProyectosSantiago"
RUTA_CARPETA_VENTAS_MENSUALES = "Ventas con ciudad 2025"
# ==============================================================================
# FUNCIONES DE AUTENTICACIÓN Y CONEXIÓN
# ==============================================================================
AUTHORITY = f"https://login.microsoftonline.com/{TENANT_ID}"
SCOPES = ["https://graph.microsoft.com/.default"]


def actualizar_archivo_trm(headers, site_id, ruta_archivo_trm, df_datos_procesados, status_placeholder):
    """
    Actualiza la hoja "Datos" del TRM.xlsx añadiendo solo nuevas filas ÚNICAS
    (evitando duplicados) y extendiendo las Tablas de Excel.
    
    INCLUYE DIAGNÓSTICOS VISUALES DE STREAMLIT para la deduplicación.
    
    VERSIÓN 4: Corrige el TypeError (len of float) en el panel de diagnóstico.
    """
    nombre_hoja_destino = "Datos"
    status_placeholder.info(f"🔄 Iniciando actualización (modo apéndice) de la hoja '{nombre_hoja_destino}'...")

    try:
        # PASO 1: Descargar el archivo
        status_placeholder.info("1/7 - Descargando archivo TRM...")
        contenido_trm_bytes = obtener_contenido_archivo_sharepoint(headers, site_id, ruta_archivo_trm)
        if contenido_trm_bytes is None: return False

        # ==============================================================================
        # PASO 2: LEER DATOS EXISTENTES Y PREPARAR DATOS NUEVOS
        # ==============================================================================
        status_placeholder.info("2/7 - Leyendo datos existentes para deduplicación...")
        
        libro_temp = openpyxl.load_workbook(io.BytesIO(contenido_trm_bytes))
        if nombre_hoja_destino not in libro_temp.sheetnames:
            status_placeholder.error(f"❌ No se encontró la hoja '{nombre_hoja_destino}'.")
            return False
        
        hoja_temp = libro_temp[nombre_hoja_destino]
        columnas_destino = [cell.value for cell in hoja_temp[1] if cell.value is not None]
        num_encabezados = len(columnas_destino)

        col_d_nombre = None
        col_aj_nombre = None
        col_ak_nombre = None

        if len(columnas_destino) > 3:
            col_d_nombre = columnas_destino[3] # Col D (índice 3)
        if len(columnas_destino) > 35:
            col_aj_nombre = columnas_destino[35] # Col AJ (índice 35)
        if len(columnas_destino) > 36:
            col_ak_nombre = columnas_destino[36] # Col AK (índice 36)
        
        cols_formula_a_ignorar = [col for col in [col_d_nombre, col_aj_nombre, col_ak_nombre] if col is not None]
        
        if cols_formula_a_ignorar:
            status_placeholder.info(f"Deduplicación ignorará columnas de fórmula: {', '.join(cols_formula_a_ignorar)}")

        df_existente = pd.read_excel(io.BytesIO(contenido_trm_bytes), sheet_name=nombre_hoja_destino, engine='openpyxl')
        df_existente.reset_index(drop=True, inplace=True)
        df_existente.columns = columnas_destino[:len(df_existente.columns)]


        status_placeholder.info(f"3/7 - Preparando {len(df_datos_procesados)} nuevos registros...")
        
        fecha_actual = datetime.now()
        dia_actual = fecha_actual.day
        anio = fecha_actual.year
        mes = fecha_actual.month
        
        if dia_actual == 1:
            if mes == 1:
                mes = 12
                anio -= 1
            else:
                mes -= 1
        
        status_placeholder.info(f"Usando fecha: Año {anio}, Mes {mes}")

        nuevas_filas_list_of_dicts = []
        for index, fila_procesada in df_datos_procesados.iterrows():
            nueva_fila_dict = {col: "" for col in columnas_destino}
            
            if len(columnas_destino) > 0: nueva_fila_dict[columnas_destino[0]] = anio
            if len(columnas_destino) > 1: nueva_fila_dict[columnas_destino[1]] = mes
            if len(columnas_destino) > 2: nueva_fila_dict[columnas_destino[2]] = "Colombia"
            
            for i, valor in enumerate(fila_procesada.values):
                col_index_destino = i + 4 
                if col_index_destino < num_encabezados:
                    nueva_fila_dict[columnas_destino[col_index_destino]] = valor
            
            nuevas_filas_list_of_dicts.append(nueva_fila_dict)

        if nuevas_filas_list_of_dicts:
            df_nuevos_mapeados = pd.DataFrame(nuevas_filas_list_of_dicts)
            df_nuevos_mapeados = df_nuevos_mapeados[columnas_destino] 
        else:
            df_nuevos_mapeados = pd.DataFrame(columns=columnas_destino)

        # ==============================================================================
        # PASO 3: DIAGNÓSTICO VISUAL DE TIPOS DE DATOS
        # ==============================================================================
        status_placeholder.info("🔍 DIAGNÓSTICO: Comparando tipos de datos...")
        
        st.write("### 📊 TIPOS DE DATOS - ARCHIVO TRM EXISTENTE (Fila 2 / Índice 0)")
        if len(df_existente) > 0:
            tipos_existente = {}
            for col in df_existente.columns:
                valor = df_existente.iloc[0][col]
                tipo = type(valor).__name__
                tipos_existente[col] = f"{tipo} | Valor: {str(valor)[:50]}"
            st.dataframe(pd.DataFrame({'Columna': list(tipos_existente.keys()), 'Tipo y Valor': list(tipos_existente.values())}))
        else:
            st.warning("⚠️ El archivo TRM existente no tiene datos (solo encabezados).")
        
        st.write("### 📊 TIPOS DE DATOS - DATOS NUEVOS MAPEADOS (Primera fila nueva / Índice 0)")
        if len(df_nuevos_mapeados) > 0:
            tipos_nuevos = {}
            for col in df_nuevos_mapeados.columns:
                valor = df_nuevos_mapeados.iloc[0][col]
                tipo = type(valor).__name__
                tipos_nuevos[col] = f"{tipo} | Valor: {str(valor)[:50]}"
            st.dataframe(pd.DataFrame({'Columna': list(tipos_nuevos.keys()), 'Tipo y Valor': list(tipos_nuevos.values())}))
        else:
            st.warning("⚠️ No hay datos nuevos para agregar.")

        st.write("### 🔍 COMPARACIÓN DE DIFERENCIAS DE TIPO")
        columnas_comunes = set(df_existente.columns) & set(df_nuevos_mapeados.columns)
        diferencias_tipo = []
        
        if len(df_existente) > 0 and len(df_nuevos_mapeados) > 0:
            for col in columnas_comunes:
                tipo_existente = type(df_existente.iloc[0][col]).__name__
                tipo_nuevo = type(df_nuevos_mapeados.iloc[0][col]).__name__
                if tipo_existente != tipo_nuevo:
                    diferencias_tipo.append({
                        'Columna': col,
                        'Tipo Existente': tipo_existente,
                        'Valor Existente': str(df_existente.iloc[0][col])[:50],
                        'Tipo Nuevo': tipo_nuevo,
                        'Valor Nuevo': str(df_nuevos_mapeados.iloc[0][col])[:50]
                    })
        
        if diferencias_tipo:
            st.error("❌ COLUMNAS CON TIPOS DE DATOS DIFERENTES:")
            st.dataframe(pd.DataFrame(diferencias_tipo))
        else:
            st.success("✅ Todos los tipos de datos coinciden en columnas comunes (basado en la primera fila).")
        
        # ==============================================================================
        # PASO 4: LÓGICA DE DEDUPLICACIÓN
        # ==============================================================================
        status_placeholder.info(f"4/7 - Ejecutando lógica de deduplicación (ignorando columnas de fórmula)...")
        
        df_existente['__source__'] = 'existente'
        df_nuevos_mapeados['__source__'] = 'nuevo'

        df_combinado = pd.concat([df_existente, df_nuevos_mapeados], ignore_index=True)
        
        df_temp_string = df_combinado.copy()
        
        cols_to_normalize = [
            col for col in df_temp_string.columns 
            if col != '__source__' and col not in cols_formula_a_ignorar
        ]

        for col in cols_to_normalize:
            try:
                if df_temp_string[col].dtype in ['float64', 'float32', 'int64', 'int32']:
                    df_temp_string[col] = df_temp_string[col].round(2)
            except:
                pass 
            
            df_temp_string[col] = (
                df_temp_string[col]
                .fillna('')
                .astype(str)
                .str.replace(r'\.0+$', '', regex=True) 
                .str.replace('None', '', regex=False)  
                .str.strip()
            )
        
        mascara_duplicados = df_temp_string.duplicated(subset=cols_to_normalize, keep='first')
        
        duplicados_encontrados_en_nuevos = mascara_duplicados[df_combinado['__source__'] == 'nuevo'].sum()
        if duplicados_encontrados_en_nuevos > 0:
            status_placeholder.warning(f"⚠️ Se encontraron {duplicados_encontrados_en_nuevos} registros nuevos que ya existían y serán omitidos.")
        else:
            status_placeholder.info("✅ No se encontraron registros duplicados en los nuevos.")

        # ==============================================================================
        # PASO 5: INVESTIGACIÓN VISUAL DE DUPLICADOS
        # ==============================================================================
        if len(df_nuevos_mapeados) > 0:
            st.write("### 🔍 INVESTIGANDO REGISTROS NO DETECTADOS COMO DUPLICADOS (EN TRM)")
            
            inicio_nuevos = len(df_existente)
            registros_nuevos_no_duplicados = sum(~mascara_duplicados[inicio_nuevos:])
            
            st.warning(f"⚠️ De {len(df_nuevos_mapeados)} registros nuevos, {registros_nuevos_no_duplicados} NO fueron detectados como duplicados y se añadirán.")
            
            if registros_nuevos_no_duplicados > 0 and duplicados_encontrados_en_nuevos > 0:
                st.info("Esto significa que ALGUNOS se detectaron y OTROS NO. Investigando diferencias...")
                
                indices_nuevos_no_detectados = [i for i in range(inicio_nuevos, len(df_combinado)) if not mascara_duplicados[i]]
                
                if indices_nuevos_no_detectados:
                    indice_problema = indices_nuevos_no_detectados[0]
                    
                    st.write(f"#### Analizando registro en índice {indice_problema} (NO detectado como duplicado)")
                    
                    col_key_nombre = None
                    if len(columnas_destino) > 10:
                        col_key_nombre = columnas_destino[10] # Columna K (Código)
                    
                    if col_key_nombre is None or col_key_nombre not in df_temp_string.columns:
                        st.warning("No se pudo identificar la columna 'Código' (K) para buscar gemelos.")
                    else:
                        codigo_buscar = df_temp_string.iloc[indice_problema][col_key_nombre]
                        st.write(f"Buscando en registros existentes con '{col_key_nombre}': **{codigo_buscar}**")
                        
                        posible_gemelo = None
                        for i in range(inicio_nuevos):
                            if df_temp_string.iloc[i][col_key_nombre] == codigo_buscar:
                                posible_gemelo = i
                                break
                        
                        if posible_gemelo is not None:
                            st.success(f"✅ Encontrado posible gemelo en índice {posible_gemelo}")
                            
                            diferencias_detalladas = []
                            for col in [c for c in df_temp_string.columns if c != '__source__']:
                                val_existente = df_temp_string.iloc[posible_gemelo][col]
                                val_nuevo = df_temp_string.iloc[indice_problema][col]
                                
                                # Convertimos a string aquí para la comparación de texto
                                val_existente_str = str(val_existente)
                                val_nuevo_str = str(val_nuevo)
                                
                                # --- INICIO DE LA CORRECCIÓN ---
                                # Normalizamos los valores de la misma forma que en la deduplicación
                                # para la comparación visual, PERO lo hacemos con str() primero
                                # para evitar el error len(float)
                                
                                if col in cols_to_normalize:
                                    # Si es una columna normalizada, aplicamos las reglas de string
                                    val_existente_str = val_existente_str.replace('.0', '').strip()
                                    val_nuevo_str = val_nuevo_str.replace('.0', '').strip()
                                # Si no, simplemente usamos la conversión a str()
                                # --- FIN DE LA CORRECCIÓN PARCIAL ---


                                if val_existente_str != val_nuevo_str:
                                    fue_ignorada = "SÍ (Fórmula)" if col in cols_formula_a_ignorar else "NO"
                                    
                                    # --- INICIO DE LA CORRECCIÓN FINAL ---
                                    # Usamos str(val_existente) y str(val_nuevo) para el len()
                                    diferencias_detalladas.append({
                                        'Columna': col,
                                        'Valor Existente': f'"{val_existente_str}" (len={len(val_existente_str)})',
                                        'Valor Nuevo': f'"{val_nuevo_str}" (len={len(val_nuevo_str)})',
                                        'Ignorada en Dedupl.': fue_ignorada
                                    })
                                    # --- FIN DE LA CORRECCIÓN FINAL ---
                            
                            if diferencias_detalladas:
                                st.error(f"❌ Encontradas {len(diferencias_detalladas)} columnas diferentes (comparando como texto):")
                                st.dataframe(pd.DataFrame(diferencias_detalladas))
                                
                                st.write("#### Valores ORIGINALES (con tipos de datos originales):")
                                diferencias_originales = []
                                for diff in diferencias_detalladas:
                                    col_name = diff['Columna']
                                    val_orig_existente = df_combinado.iloc[posible_gemelo][col_name]
                                    val_orig_nuevo = df_combinado.iloc[indice_problema][col_name]
                                    
                                    diferencias_originales.append({
                                        'Columna': col_name,
                                        'Valor Existente': val_orig_existente,
                                        'Tipo Existente': type(val_orig_existente).__name__,
                                        'Valor Nuevo': val_orig_nuevo,
                                        'Tipo Nuevo': type(val_orig_nuevo).__name__
                                    })
                                st.dataframe(pd.DataFrame(diferencias_originales))
                                
                            else:
                                st.success("✅ No se encontraron diferencias en la comparación de texto.")
                        else:
                            st.warning(f"⚠️ No se encontró un registro existente con '{col_key_nombre}' = {codigo_buscar}. Este registro es genuinamente nuevo.")

        # ==============================================================================
        # PASO 6: AÑADIR LAS FILAS ÚNICAS Y SUBIR
        # ==============================================================================
        
        df_combinado_filtrado = df_combinado[~mascara_duplicados]
        df_filas_a_anadir = df_combinado_filtrado[df_combinado_filtrado['__source__'] == 'nuevo']
        df_filas_a_anadir = df_filas_a_anadir.drop(columns=['__source__'])

        if df_filas_a_anadir.empty:
            status_placeholder.success("✅ No se encontraron registros nuevos para añadir. El archivo TRM ya está actualizado.")
            return True

        status_placeholder.info(f"5/7 - Añadiendo {len(df_filas_a_anadir)} nuevos registros únicos...")

        lista_nuevas_filas_final = [list(row) for row in df_filas_a_anadir.itertuples(index=False, name=None)]

        libro = openpyxl.load_workbook(io.BytesIO(contenido_trm_bytes))
        hoja = libro[nombre_hoja_destino]
        
        tabla = None
        if hoja.tables:
            nombre_tabla = list(hoja.tables.keys())[0]
            tabla = hoja.tables[nombre_tabla]
            status_placeholder.info(f"✅ Tabla encontrada: '{nombre_tabla}' con rango {tabla.ref}")
        else:
            status_placeholder.warning("⚠️ No se encontró ninguna Tabla de Excel.")

        for fila in lista_nuevas_filas_final:
            hoja.append(fila)
        
        if tabla:
            rango_actual = tabla.ref
            inicio_rango = rango_actual.split(':')[0]
            columna_final = rango_actual.split(':')[1].rstrip('0123456789')
            nueva_fila_final = hoja.max_row
            nuevo_rango = f"{inicio_rango}:{columna_final}{nueva_fila_final}"
            tabla.ref = nuevo_rango
            status_placeholder.info(f"✅ Rango de la Tabla extendido de {rango_actual} a {nuevo_rango}")

        # ==============================================================================
        # PASO 7: INYECTAR FÓRMULAS
        # ==============================================================================
        status_placeholder.info("6/7 - Agregando fórmulas a las nuevas filas...")

        col_comercial_idx = 4
        col_vendedor_idx = 18
        
        num_nuevas_filas = len(lista_nuevas_filas_final)
        primera_fila_nueva = hoja.max_row - num_nuevas_filas + 1
        
        for r_idx in range(primera_fila_nueva, hoja.max_row + 1):
            celda_comercial = hoja.cell(row=r_idx, column=col_comercial_idx)
            celda_comercial.value = f'=IFERROR(VLOOKUP(R{r_idx},vendedor!$B:$C,2,FALSE),"")'
        
        status_placeholder.info(f"✅ Fórmula agregada a columna D en {num_nuevas_filas} nuevas filas")

        col_aj_idx = 36
        col_ak_idx = 37
        
        for r_idx in range(primera_fila_nueva, hoja.max_row + 1):
            celda_aj = hoja.cell(row=r_idx, column=col_aj_idx)
            celda_aj.value = f'=IFERROR(1-(AH{r_idx}/W{r_idx}),0)'
            
            celda_ak = hoja.cell(row=r_idx, column=col_ak_idx)
            celda_ak.value = f'=W{r_idx}-AH{r_idx}'
        
        status_placeholder.info(f"✅ Fórmulas agregadas a columnas AJ y AK en {num_nuevas_filas} nuevas filas")

        # ==============================================================================
        # PASO 8: GUARDAR Y SUBIR
        # ==============================================================================
        status_placeholder.info("7/7 - Guardando y subiendo archivo final...")
        
        output = io.BytesIO()
        libro.save(output)
        
        endpoint_put = f"https://graph.microsoft.com/v1.0/sites/{site_id}/drive/root:/{ruta_archivo_trm}:/content"
        response_put = requests.put(endpoint_put, data=output.getvalue(), headers=headers)
        response_put.raise_for_status()

        status_placeholder.success(f"✅ ¡Archivo TRM actualizado! Se añadieron {num_nuevas_filas} registros nuevos y únicos.")
        return True

    except Exception as e:
        status_placeholder.error(f"❌ Falló la actualización del archivo TRM. Error: {e}")
        import traceback
        status_placeholder.error(f"Detalles del error: {traceback.format_exc()}")
        return False


def validar_respuesta_sharepoint(response, nombre_archivo):
    """
    Valida que la respuesta de SharePoint sea correcta y contenga un archivo Excel
    """
    #st.info(f"🔍 Validando respuesta para: {nombre_archivo}")
    
    # 1. Verificar código de estado HTTP
    #st.write(f"📊 Código HTTP: {response.status_code}")
    
    if response.status_code != 200:
        #st.error(f"❌ Error HTTP {response.status_code}")
        try:
            error_json = response.json()
            #st.json(error_json)
        except:
            st.error(f"Texto de respuesta: {response.text[:500]}...")
        return False, "Error HTTP"
    
    # 2. Verificar el tamaño del contenido
    content_length = len(response.content)
    #st.write(f"📏 Tamaño del archivo descargado: {content_length:,} bytes")
    
    if content_length == 0:
        st.error("❌ El archivo está vacío (0 bytes)")
        return False, "Archivo vacío"
    
    if content_length < 100:  # Un Excel válido debe tener al menos algunos cientos de bytes
        #st.warning("⚠️ El archivo es muy pequeño para ser un Excel válido")
        #st.write(f"Contenido recibido: {response.content}")
        return False, "Archivo muy pequeño"
    
    # 3. Verificar el Content-Type si está disponible
    content_type = response.headers.get('Content-Type', 'No especificado')
    #st.write(f"📋 Content-Type: {content_type}")
    
    # 4. Verificar las primeras bytes para asegurar que es un archivo Excel
    primeros_bytes = response.content[:20]
    #st.write(f"🔢 Primeros 20 bytes (hex): {primeros_bytes.hex()}")
    
    # Un archivo Excel (.xlsx) debe comenzar con la signature de ZIP: "PK"
    if not response.content.startswith(b'PK'):
        #st.error("❌ El archivo no tiene la signature de un archivo ZIP/Excel válido")
        #st.error("Los archivos .xlsx deben comenzar con 'PK' (signature de ZIP)")
        
        # Mostrar el inicio del contenido como texto para debug
        try:
            inicio_texto = response.content[:200].decode('utf-8', errors='ignore')
            #st.error(f"Inicio del contenido como texto: {inicio_texto}")
        except:
            st.error("No se pudo decodificar el inicio del contenido como texto")
        
        return False, "Signature inválida"
    
    #st.success("✅ El archivo parece ser un Excel válido")
    return True, "Válido"

def obtener_contenido_archivo_sharepoint(headers, site_id, ruta_archivo):
    """
    Descarga un archivo específico de SharePoint con validaciones completas
    """
    #st.info(f"📥 Descargando archivo: {ruta_archivo}")
    
    # Construir el endpoint
    endpoint_get = f"https://graph.microsoft.com/v1.0/sites/{site_id}/drive/root:/{ruta_archivo}:/content"
    #st.write(f"🔗 Endpoint: {endpoint_get}")
    
    try:
        # Realizar la petición
        response_get = requests.get(endpoint_get, headers=headers)
        
        # Validar la respuesta
        es_valido, mensaje = validar_respuesta_sharepoint(response_get, ruta_archivo.split('/')[-1])
        
        if not es_valido:
            #st.error(f"❌ Validación falló: {mensaje}")
            return None
        
        return response_get.content
        
    except requests.exceptions.RequestException as e:
        #st.error(f"❌ Error de red al descargar el archivo: {e}")
        return None
    except Exception as e:
        #st.error(f"❌ Error inesperado: {e}")
        return None

def verificar_archivo_existe_sharepoint(headers, site_id, ruta_archivo):
    """
    Verifica si un archivo existe y obtiene sus metadatos antes de descargarlo
    """
    #st.info(f"🔍 Verificando existencia de: {ruta_archivo}")
    
    # Endpoint para obtener metadatos del archivo (sin descargar el contenido)
    endpoint_metadata = f"https://graph.microsoft.com/v1.0/sites/{site_id}/drive/root:/{ruta_archivo}"
    
    try:
        response = requests.get(endpoint_metadata, headers=headers)
        
        if response.status_code == 200:
            metadata = response.json()
            
            nombre = metadata.get('name', 'Sin nombre')
            tamano = metadata.get('size', 0)
            tipo = metadata.get('file', {}).get('mimeType', 'No especificado')
            modificado = metadata.get('lastModifiedDateTime', 'No especificado')
            
            #st.success(f"✅ Archivo encontrado: {nombre}")
            #st.write(f"📏 Tamaño: {tamano:,} bytes")
            #st.write(f"📋 Tipo MIME: {tipo}")
            #st.write(f"📅 Última modificación: {modificado}")
            
            # Verificar que sea realmente un archivo Excel
            if tipo and 'spreadsheet' not in tipo.lower() and 'excel' not in tipo.lower():
                st.warning(f"⚠️ Advertencia: El tipo MIME '{tipo}' no parece ser un Excel")
            
            return True, metadata
        else:
            #st.error(f"❌ Archivo no encontrado. HTTP {response.status_code}")
            try:
                error_json = response.json()
                #st.json(error_json)
            except:
                st.error(f"Respuesta: {response.text}")
            return False, None
            
    except Exception as e:
        #t.error(f"❌ Error al verificar archivo: {e}")
        return False, None


def get_access_token(status_placeholder):
    #status_placeholder.info("⚙️ Paso 2/5: Autenticando con Microsoft...")
    app = ConfidentialClientApplication(
        client_id=CLIENT_ID,
        authority=AUTHORITY,
        client_credential=CLIENT_SECRET
    )
    result = app.acquire_token_for_client(scopes=SCOPES)
    if "access_token" in result:
        #st.success("✅ Token de acceso obtenido con éxito.")
        return result['access_token']
    else:
        #st.error(f"Error al obtener token: {result.get('error_description')}")
        return None

def get_sharepoint_site_id(access_token):
    headers = {'Authorization': f'Bearer {access_token}'}
    site_url = f"https://graph.microsoft.com/v1.0/sites/{SHAREPOINT_HOSTNAME}:/sites/{SITE_NAME}"
    try:
        response = requests.get(site_url, headers=headers)
        response.raise_for_status()
        site_id = response.json().get('id')
        #st.success(f"✅ Conexión exitosa con el sitio SharePoint: '{SITE_NAME}'")
        return site_id
    except requests.exceptions.RequestException as e:
        #st.error(f"Error al obtener site_id: {e.response.text}")
        return None

def encontrar_archivo_del_mes(headers, site_id, ruta_carpeta, status_placeholder):
    """
    Busca dentro de una CARPETA específica y devuelve la RUTA COMPLETA del archivo del mes.
    """
    try:
        # Meses en español con diferentes variaciones
        fecha_actual = datetime.now()
        mes_numero = fecha_actual.month
        
        # Diferentes patrones que podría tener el archivo
        patrones_busqueda = [
            f"{mes_numero}. ",  # "9. " para septiembre
            "Noviembre",       # Nombre completo del mes
            "noviembre",       # Minúscula
            f"{mes_numero:02d}",# "09" con cero delante
        ]
        
        #st.info(f"🔍 Buscando archivo del mes {mes_numero} (Septiembre) en: '{ruta_carpeta}'")
        #st.write(f"Patrones de búsqueda: {patrones_busqueda}")
        
        # Primero, listar TODOS los archivos en la carpeta
        #st.write("📂 Listando todos los archivos disponibles:")
        endpoint_children = f"https://graph.microsoft.com/v1.0/sites/{site_id}/drive/root:/{ruta_carpeta}:/children"
        response_list = requests.get(endpoint_children, headers=headers)
        
        if response_list.status_code == 200:
            todos_archivos = response_list.json().get('value', [])
            
            #st.write(f"📊 Total de archivos en la carpeta: {len(todos_archivos)}")
            
            # Mostrar todos los archivos para debug
            for item in todos_archivos:
                if not item.get('folder'):  # Solo archivos, no carpetas
                    nombre = item.get('name', '')
                    tamaño = item.get('size', 0)
                    #st.write(f"📄 {nombre} ({tamaño:,} bytes)")
            
            # Buscar el archivo que coincida con los patrones
            archivos_candidatos = []
            
            for item in todos_archivos:
                if item.get('folder'):  # Saltar carpetas
                    continue
                    
                nombre_archivo = item.get('name', '').lower()
                
                # Verificar cada patrón
                for patron in patrones_busqueda:
                    if patron.lower() in nombre_archivo:
                        archivos_candidatos.append({
                            'nombre_original': item.get('name'),
                            'ruta_completa': f"{ruta_carpeta}/{item.get('name')}",
                            'tamaño': item.get('size', 0),
                            'patron_encontrado': patron
                        })
                        break  # Salir del loop de patrones una vez encontrado
            
            if archivos_candidatos:
                #st.success(f"✅ Encontrados {len(archivos_candidatos)} archivos candidatos:")
                
                #for i, candidato in enumerate(archivos_candidatos):
                    #st.write(f"{i+1}. **{candidato['nombre_original']}** ({candidato['tamaño']:,} bytes) - Patrón: '{candidato['patron_encontrado']}'")
                
                # Seleccionar el primer candidato (o puedes agregar lógica más sofisticada)
                archivo_seleccionado = archivos_candidatos[0]
                #st.success(f"🎯 Archivo seleccionado: **{archivo_seleccionado['nombre_original']}**")
                
                return archivo_seleccionado['ruta_completa']
            else:
                #st.warning(f"⚠️ No se encontraron archivos que coincidan con los patrones para el mes {mes_numero}")
                
                # Mostrar sugerencia
                #st.info("💡 Archivos disponibles que podrían ser relevantes:")
                for item in todos_archivos:
                    if not item.get('folder'):
                        nombre = item.get('name', '')
                        if any(char.isdigit() for char in nombre):  # Si contiene números
                            st.write(f"🤔 {nombre}")
                
                return None
        else:
            st.error(f"❌ No se pudo listar el contenido de la carpeta. HTTP {response_list.status_code}")
            return None
            
    except requests.exceptions.RequestException as e:
        st.error(f"Error de conexión al buscar el archivo del mes: {e.response.text if e.response else e}")
        return None
    except Exception as e:
        st.error(f"Error inesperado durante la búsqueda del mes: {e}")
        return None

def agregar_datos_a_excel_sharepoint(headers, site_id, ruta_archivo, df_nuevos_datos, status_placeholder):
    """
    Agrega datos a la primera hoja de un archivo Excel en SharePoint,
    preservando fórmulas, formatos y otras hojas, y eliminando duplicados
    mediante comparación temporal de strings sin modificar los tipos de datos originales.
    """
    try:
        # PASO 1: Descargar el archivo existente con validaciones
        contenido_bytes = obtener_contenido_archivo_sharepoint(headers, site_id, ruta_archivo)
        if contenido_bytes is None:
            return False

        contenido_en_memoria = io.BytesIO(contenido_bytes)

        # PASO 2: Cargar el libro de trabajo completo con openpyxl
        libro = openpyxl.load_workbook(contenido_en_memoria)
        
        nombre_hoja_destino = libro.sheetnames[0]
        hoja = libro[nombre_hoja_destino]
        
        # Leer los datos de esa hoja en un DataFrame
        df_existente = pd.read_excel(io.BytesIO(contenido_bytes), sheet_name=nombre_hoja_destino, engine='openpyxl')
        df_existente.reset_index(drop=True, inplace=True)

        # ====== DIAGNÓSTICO: COMPARAR TIPOS DE DATOS ======
        status_placeholder.info("🔍 DIAGNÓSTICO: Comparando tipos de datos...")
        
        st.write("### 📊 TIPOS DE DATOS - ARCHIVO EXISTENTE (Fila 2 / Índice 0)")
        if len(df_existente) > 0:
            st.write("**Tipos de datos por columna:**")
            tipos_existente = {}
            for col in df_existente.columns:
                valor = df_existente.iloc[0][col]
                tipo = type(valor).__name__
                tipos_existente[col] = f"{tipo} | Valor: {valor}"
            
            # Mostrar en formato tabla
            st.dataframe(pd.DataFrame({
                'Columna': list(tipos_existente.keys()),
                'Tipo y Valor': list(tipos_existente.values())
            }))
        else:
            st.warning("⚠️ El archivo existente no tiene datos")
        
        st.write("### 📊 TIPOS DE DATOS - DATOS NUEVOS (Primera fila nueva / Índice 0)")
        if len(df_nuevos_datos) > 0:
            st.write("**Tipos de datos por columna:**")
            tipos_nuevos = {}
            for col in df_nuevos_datos.columns:
                valor = df_nuevos_datos.iloc[0][col]
                tipo = type(valor).__name__
                tipos_nuevos[col] = f"{tipo} | Valor: {valor}"
            
            # Mostrar en formato tabla
            st.dataframe(pd.DataFrame({
                'Columna': list(tipos_nuevos.keys()),
                'Tipo y Valor': list(tipos_nuevos.values())
            }))
        else:
            st.warning("⚠️ No hay datos nuevos para agregar")
        
        st.write("### 🔍 COMPARACIÓN DE DIFERENCIAS")
        # Comparar columnas comunes
        columnas_comunes = set(df_existente.columns) & set(df_nuevos_datos.columns)
        diferencias_tipo = []
        
        for col in columnas_comunes:
            if len(df_existente) > 0 and len(df_nuevos_datos) > 0:
                tipo_existente = type(df_existente.iloc[0][col]).__name__
                tipo_nuevo = type(df_nuevos_datos.iloc[0][col]).__name__
                
                if tipo_existente != tipo_nuevo:
                    valor_existente = df_existente.iloc[0][col]
                    valor_nuevo = df_nuevos_datos.iloc[0][col]
                    diferencias_tipo.append({
                        'Columna': col,
                        'Tipo Existente': tipo_existente,
                        'Valor Existente': valor_existente,
                        'Tipo Nuevo': tipo_nuevo,
                        'Valor Nuevo': valor_nuevo
                    })
        
        if diferencias_tipo:
            st.error("❌ COLUMNAS CON TIPOS DE DATOS DIFERENTES:")
            st.dataframe(pd.DataFrame(diferencias_tipo))
        else:
            st.success("✅ Todos los tipos de datos coinciden en columnas comunes")
        
        # ====== FIN DIAGNÓSTICO ======

        # PASO 3: Combinar datos y eliminar duplicados
        status_placeholder.info("3/4 - Combinando datos nuevos y existentes...")
        df_combinado = pd.concat([df_existente, df_nuevos_datos], ignore_index=True)
        
        # ====== DETECCIÓN DE DUPLICADOS CON COMPARACIÓN TEMPORAL STRING ======
        filas_antes = len(df_combinado)
        status_placeholder.info(f"📊 Total de filas antes de verificar duplicados: {filas_antes}")
        
        # CREAR COPIA TEMPORAL DE TODO EL DATAFRAME PARA COMPARACIÓN
        df_temp_string = df_combinado.copy()
        
        status_placeholder.info(f"🔍 Convirtiendo {len(df_temp_string.columns)} columnas a string para comparación...")
        
        # Convertir TODAS las columnas a string para comparación uniforme
        for col in df_temp_string.columns:
            # 1. Rellenar NaN/None con string vacío
            # 2. Si es numérico, redondear a 2 decimales para evitar diferencias de precisión
            # 3. Convertir a string
            # 4. Eliminar el .0 al final de números flotantes
            # 5. Reemplazar "None" con ""
            # 6. Limpiar espacios
            
            # Intentar redondear si es numérico
            try:
                # Si la columna es numérica, redondear a 2 decimales
                if df_temp_string[col].dtype in ['float64', 'float32', 'int64', 'int32']:
                    df_temp_string[col] = df_temp_string[col].round(2)
            except:
                pass  # Si no es numérico, continuar
            
            df_temp_string[col] = (
                df_temp_string[col]
                .fillna('')
                .astype(str)
                .str.replace(r'\.0+$', '', regex=True)  # "2.0" → "2"
                .str.replace('None', '', regex=False)    # "None" → ""
                .str.strip()
            )
        
        status_placeholder.info("✅ Todas las columnas convertidas a string (redondeadas y normalizadas).")
        
        # IDENTIFICAR duplicados usando la versión temporal en string
        # Esto compara TODAS las columnas de cada registro
        mascara_duplicados = df_temp_string.duplicated(keep='first')
        
        # Contar duplicados encontrados
        duplicados_encontrados = mascara_duplicados.sum()
        
        if duplicados_encontrados > 0:
            status_placeholder.warning(
                f"⚠️ Se encontraron {duplicados_encontrados} registros duplicados que serán omitidos."
            )
            
            # Opcional: Mostrar algunos ejemplos de duplicados para debug
            indices_duplicados = df_combinado[mascara_duplicados].index.tolist()[:3]
            if indices_duplicados:
                status_placeholder.info(f"📋 Ejemplos de índices de filas duplicadas: {indices_duplicados}")
        else:
            status_placeholder.success("✅ No se encontraron registros duplicados.")
        
        # ====== INVESTIGAR REGISTROS NO DETECTADOS ======
        if len(df_nuevos_datos) > 0:
            st.write("### 🔍 INVESTIGANDO REGISTROS NO DETECTADOS COMO DUPLICADOS")
            
            # Los nuevos datos están al final del df_combinado
            inicio_nuevos = len(df_existente)
            
            # Ver cuántos de los nuevos NO fueron marcados como duplicados
            registros_nuevos_no_duplicados = sum(~mascara_duplicados[inicio_nuevos:])
            
            st.warning(f"⚠️ De {len(df_nuevos_datos)} registros nuevos, {registros_nuevos_no_duplicados} NO fueron detectados como duplicados")
            
            if registros_nuevos_no_duplicados > 0 and registros_nuevos_no_duplicados < len(df_nuevos_datos):
                st.write("Esto significa que ALGUNOS se detectaron y OTROS NO. Investigando diferencias...")
                
                # Obtener los índices de registros nuevos que NO fueron detectados como duplicados
                indices_nuevos_no_detectados = []
                for i in range(inicio_nuevos, len(df_combinado)):
                    if not mascara_duplicados[i]:
                        indices_nuevos_no_detectados.append(i)
                
                if indices_nuevos_no_detectados:
                    # Tomar el primer registro nuevo que NO se detectó como duplicado
                    indice_problema = indices_nuevos_no_detectados[0]
                    
                    st.write(f"#### Analizando registro en índice {indice_problema} (NO detectado como duplicado)")
                    
                    # Buscar registros existentes que tengan el mismo "Código" (columna clave)
                    if 'Código' in df_temp_string.columns:
                        codigo_buscar = df_temp_string.iloc[indice_problema]['Código']
                        
                        st.write(f"Buscando en registros existentes con Código: **{codigo_buscar}**")
                        
                        # Buscar en los registros existentes (antes de inicio_nuevos)
                        posible_gemelo = None
                        for i in range(inicio_nuevos):
                            if df_temp_string.iloc[i]['Código'] == codigo_buscar:
                                posible_gemelo = i
                                break
                        
                        if posible_gemelo is not None:
                            st.success(f"✅ Encontrado posible gemelo en índice {posible_gemelo}")
                            
                            # Comparar TODAS las columnas entre estos dos registros
                            diferencias_detalladas = []
                            for col in df_temp_string.columns:
                                val_existente = df_temp_string.iloc[posible_gemelo][col]
                                val_nuevo = df_temp_string.iloc[indice_problema][col]
                                
                                if val_existente != val_nuevo:
                                    # Mostrar también el tipo y longitud para debugging
                                    diferencias_detalladas.append({
                                        'Columna': col,
                                        'Valor Existente (ya string)': f'"{val_existente}" (len={len(val_existente)})',
                                        'Valor Nuevo (ya string)': f'"{val_nuevo}" (len={len(val_nuevo)})',
                                        'Son iguales?': 'NO ❌'
                                    })
                            
                            if diferencias_detalladas:
                                st.error(f"❌ Encontradas {len(diferencias_detalladas)} columnas diferentes:")
                                st.dataframe(pd.DataFrame(diferencias_detalladas))
                                
                                # Mostrar también los valores ORIGINALES (antes de convertir a string)
                                st.write("#### Valores ORIGINALES (con tipos de datos originales):")
                                diferencias_originales = []
                                for col in df_combinado.columns:
                                    val_orig_existente = df_combinado.iloc[posible_gemelo][col]
                                    val_orig_nuevo = df_combinado.iloc[indice_problema][col]
                                    tipo_existente = type(val_orig_existente).__name__
                                    tipo_nuevo = type(val_orig_nuevo).__name__
                                    
                                    if str(val_orig_existente) != str(val_orig_nuevo):
                                        diferencias_originales.append({
                                            'Columna': col,
                                            'Valor Existente': val_orig_existente,
                                            'Tipo Existente': tipo_existente,
                                            'Valor Nuevo': val_orig_nuevo,
                                            'Tipo Nuevo': tipo_nuevo
                                        })
                                
                                if diferencias_originales:
                                    st.dataframe(pd.DataFrame(diferencias_originales))
                            else:
                                st.success("✅ Todos los valores son iguales (esto NO debería pasar)")
                        else:
                            st.warning(f"⚠️ No se encontró un registro existente con Código {codigo_buscar}")
        # ====== FIN INVESTIGACIÓN ======
        
        # FILTRAR el DataFrame ORIGINAL (con tipos de datos originales intactos)
        # usando la máscara de duplicados identificada
        df_sin_duplicados = df_combinado[~mascara_duplicados].copy()
        
        filas_despues = len(df_sin_duplicados)
        status_placeholder.success(f"✅ Filas finales después de eliminar duplicados: {filas_despues}")
        
        # ====== FIN DETECCIÓN DE DUPLICADOS ======
        
        # Limpiar columnas "Unnamed"
        cols_a_eliminar = [col for col in df_sin_duplicados.columns if 'Unnamed:' in str(col)]
        if cols_a_eliminar:
            df_sin_duplicados.drop(columns=cols_a_eliminar, inplace=True)
            status_placeholder.info("🧹 Columnas 'Unnamed:' eliminadas.")

        # PASO 4: Escribir los datos actualizados de vuelta a la hoja
        status_placeholder.info("4/4 - Escribiendo datos y subiendo el archivo final...")
        
        # Usar el DataFrame limpio (sin duplicados, pero con tipos originales)
        df_combinado = df_sin_duplicados
        hoja = libro[nombre_hoja_destino]
        
        # Borrar datos antiguos de la hoja (excepto encabezados)
        for r in range(hoja.max_row, 1, -1):
            hoja.delete_rows(r)
            
        from openpyxl.utils.dataframe import dataframe_to_rows    
        
        # Escribir el contenido del DataFrame combinado en la hoja
        for r_idx, row in enumerate(dataframe_to_rows(df_sin_duplicados, index=False, header=False), 2):
            for c_idx, value in enumerate(row, 1):
                hoja.cell(row=r_idx, column=c_idx, value=value)
        
        # Guardar el libro modificado en memoria
        output = io.BytesIO()
        libro.save(output)
        
        # Subir el archivo final
        endpoint_put = f"https://graph.microsoft.com/v1.0/sites/{site_id}/drive/root:/{ruta_archivo}:/content"
        response_put = requests.put(endpoint_put, data=output.getvalue(), headers=headers)
        response_put.raise_for_status()

        status_placeholder.success(f"✅ ¡Archivo '{ruta_archivo.split('/')[-1]}' actualizado preservando su formato!")
        return True

    except Exception as e:
        status_placeholder.error(f"❌ Falló la actualización del archivo. Error: {e}")
        import traceback
        status_placeholder.error(f"Detalles del error: {traceback.format_exc()}")
        return False
    
    
def listar_archivos_en_carpeta(headers, site_id, ruta_carpeta):
    """
    Lista todos los archivos en una carpeta para debug
    """
    #st.info(f"📂 Explorando carpeta: {ruta_carpeta}")
    
    endpoint = f"https://graph.microsoft.com/v1.0/sites/{site_id}/drive/root:/{ruta_carpeta}:/children"
    
    try:
        response = requests.get(endpoint, headers=headers)
        if response.status_code == 200:
            items = response.json().get('value', [])
            
            #st.write(f"📊 Encontrados {len(items)} elementos:")
            for item in items:
                tipo = "📁" if item.get('folder') else "📄"
                nombre = item.get('name', 'Sin nombre')
                tamano = item.get('size', 0)
                #st.write(f"{tipo} {nombre} ({tamano:,} bytes)")
        else:
            st.error(f"❌ No se pudo listar la carpeta. HTTP {response.status_code}")
    except Exception as e:
        st.error(f"❌ Error: {e}")
    
    
# --- Función Principal de Procesamiento ---
def procesar_excel_para_streamlit(uploaded_file, status_placeholder):
    """
    Procesa el archivo de Excel subido:
    - Ignora las primeras 7 filas al cargar el archivo (asumiendo que los encabezados están en la fila 8).
    - Elimina filas con 'Tipo clasificación' vacío.
    - Elimina columnas no deseadas.
    - Actualiza la columna 'Total'.
    - Rellena 'Tasa de cambio' con TRM de API bajo condiciones específicas.

    Args:
        uploaded_file (streamlit.UploadedFile): El archivo Excel subido por el usuario.

    Returns:
        pandas.DataFrame or None: El DataFrame procesado o None si hay un error.
    """
    try:
        # Usar skiprows para que Pandas lea el encabezado correcto
        df = pd.read_excel(uploaded_file, skiprows=7) # La fila 8 (índice 7) se toma como encabezado

        # Verifica si el DataFrame tiene columnas después de skiprows.
        if df.empty or df.columns.empty:
            st.error("Parece que el archivo no tiene datos o encabezados después de saltar las primeras 7 filas. Por favor, verifica el formato del archivo.")
            return None

        st.info(f"Archivo cargado exitosamente. Se saltaron las primeras 7 filas. Filas iniciales (después de saltar): **{len(df)}**.")

        df_procesado = df.copy()
        
        # --- FUNCIÓN DE LIMPIEZA SIMPLE ---
        def convertir_a_numero_limpiando_comas(columna):
            if not pd.api.types.is_string_dtype(columna):
                columna = columna.astype(str)
            columna_limpia = columna.str.replace(',', '', regex=False)
            return pd.to_numeric(columna_limpia, errors='coerce')

        if 'Identificación Vendedor' in df_procesado.columns:
            # Crear la nueva columna 'Vendedor' con los datos de la original
            df_procesado['Vendedor'] = df_procesado['Identificación Vendedor']
            st.success("✅ Columna 'Vendedor' creada con éxito.")
        else:
            # Si la columna original no existe, crear 'Vendedor' como una columna vacía
            st.warning("⚠️ No se encontró la columna 'Identificación Vendedor'. Se creará una columna 'Vendedor' vacía.")
            df_procesado['Vendedor'] = ''

        # Columnas a eliminar predefinidas
        nombres_columnas_a_eliminar = [
            "Sucursal",
            "Centro costo",
            "Fecha creación",
            "Fecha modificación",
            "Correo electrónico",
            "Tipo de registro",
            "Referencia fábrica",
            "Bodega",
            #"Identificación Vendedor",
            "Nombre vendedor",
            "Valor desc.",
            "Base AIU",
            "Impuesto cargo",
            "Valor Impuesto Cargo",
            "Impuesto Cargo 2",
            "Valor Impuesto Cargo 2",
            "Impuesto retención",
            "Valor Impuesto Retención",
            "Base retención (ICA/IVA)",
            "Cargo en totales",
            "Descuento en totales",
            "Moneda",
            "Forma pago",
            "Fecha vencimiento",
            "Nombre contacto"
        ]

        #df_procesado = df.copy()
        
        #Extraer códigos de Línea y Sublínea desde "Referencia fábrica"
        if "Referencia fábrica" in df_procesado.columns:
            st.info("Extrayendo códigos de Línea y Sublínea desde 'Referencia fábrica'...")
            
            # Convertir a string para poder usar regex
            df_procesado['Referencia fábrica'] = df_procesado['Referencia fábrica'].astype(str)
            
            # Extraer código de línea (entre paréntesis) - TODO el contenido
            df_procesado['Línea'] = df_procesado['Referencia fábrica'].str.extract(r'\(([^)]+)\)', expand=False)
            
            # Extraer código de sublínea (entre llaves) - TODO el contenido
            df_procesado['Sublínea'] = df_procesado['Referencia fábrica'].str.extract(r'\{([^}]+)\}', expand=False)
            
            # Reemplazar NaN con string vacío
            df_procesado['Línea'].fillna('', inplace=True)
            df_procesado['Sublínea'].fillna('', inplace=True)
            
            st.success(f"Códigos extraídos - Líneas: {df_procesado['Línea'].ne('').sum()}, Sublíneas: {df_procesado['Sublínea'].ne('').sum()}")
        else:
            st.warning("No se encontró la columna 'Referencia fábrica'.")
            df_procesado['Línea'] = ''
            df_procesado['Sublínea'] = ''
            
        if "Observaciones" in df_procesado.columns:
            st.info("Extrayendo Clasificación Producto desde 'Observaciones'...")
            
            df_procesado['Observaciones'] = df_procesado['Observaciones'].astype(str)
            
            # Extraer contenido entre comillas dobles
            df_procesado['Clasificación Producto'] = df_procesado['Observaciones'].str.extract(r'"([^"]+)"', expand=False)
            
            # Reemplazar NaN con string vacío
            df_procesado['Clasificación Producto'].fillna('', inplace=True)
            
            clasificaciones_encontradas = df_procesado['Clasificación Producto'].ne('').sum()
            st.success(f"Clasificaciones de producto extraídas: {clasificaciones_encontradas}")
        else:
            st.warning("No se encontró la columna 'Observaciones'.")
            df_procesado['Clasificación Producto'] = ''
        
        
        # 1. Eliminar filas donde "Tipo clasificación" esté vacío/NaN
        if "Tipo clasificación" in df_procesado.columns:
            filas_antes_eliminacion = len(df_procesado)
            df_procesado.dropna(subset=["Tipo clasificación"], inplace=True)
            filas_despues_eliminacion = len(df_procesado)
            st.success(f"Filas con 'Tipo clasificación' vacío eliminadas: **{filas_antes_eliminacion - filas_despues_eliminacion}**. Filas restantes: **{filas_despues_eliminacion}**.")
        else:
            st.warning("La columna **'Tipo clasificación'** no se encontró. No se eliminaron filas vacías.")

        # 2. Eliminar columnas especificadas
        columnas_existentes_para_eliminar = [col for col in nombres_columnas_a_eliminar if col in df_procesado.columns]
        columnas_no_existentes_para_eliminar = [col for col in nombres_columnas_a_eliminar if col not in df_procesado.columns]

        if columnas_existentes_para_eliminar:
            df_procesado.drop(columns=columnas_existentes_para_eliminar, inplace=True)
            st.success(f"Columnas eliminadas: **{', '.join(columnas_existentes_para_eliminar)}**.")
        else:
            st.info("Ninguna de las columnas especificadas para eliminar se encontró. No se eliminaron columnas.")

        if columnas_no_existentes_para_eliminar:
            st.warning(f"Advertencia: Las siguientes columnas especificadas para eliminación no se encontraron: **{', '.join(columnas_no_existentes_para_eliminar)}**.")

        # 3. Actualizar la columna "Total" existente
        if "Cantidad" in df_procesado.columns and "Valor unitario" in df_procesado.columns and "Total" in df_procesado.columns:
            df_procesado["Cantidad"] = pd.to_numeric(df_procesado["Cantidad"], errors='coerce')
            df_procesado["Valor unitario"] = pd.to_numeric(df_procesado["Valor unitario"], errors='coerce')
            df_procesado["Total"] = df_procesado["Cantidad"] * df_procesado["Valor unitario"]
            df_procesado["Total"] = df_procesado["Total"].fillna(0)
            st.success("La columna **'Total'** ha sido actualizada con el cálculo **'Cantidad * Valor unitario'**.")
        else:
            st.warning("Advertencia: No se pudieron encontrar las columnas **'Cantidad'**, **'Valor unitario'** y/o **'Total'**. La columna **'Total'** no se actualizó.")

        # 4. Crear y posicionar la nueva columna "Numero comprobante"
        columnas_necesarias = ['Número comprobante', 'Consecutivo', 'Factura proveedor']
        if all(col in df_procesado.columns for col in columnas_necesarias):
            # Definir las condiciones
            conditions = [
                df_procesado['Número comprobante'] == 'FV-1',
                df_procesado['Número comprobante'] == 'FV-2'
            ]
            
            # Definir los valores a asignar para cada condición
            choices = [
                'FLE-' + df_procesado['Consecutivo'].astype('Int64').astype(str),
                'FSE-' + df_procesado['Consecutivo'].astype('Int64').astype(str)
            ]
            
            # Usar np.select para crear los valores de la nueva columna
            # El valor por defecto será un texto vacío ''
            valores_nueva_columna = np.select(conditions, choices, default='')
            
            # Encontrar la posición de la columna "Factura proveedor" para insertar antes
            posicion_insercion = df_procesado.columns.get_loc('Factura proveedor')
            
            # Insertar la nueva columna en la posición encontrada
            df_procesado.insert(posicion_insercion, 'Numero comprobante', valores_nueva_columna)
            
            st.success("Se ha creado y llenado la nueva columna **'Numero comprobante'**.")
            
        else:
            st.warning("Advertencia: No se encontraron las columnas necesarias ('Número comprobante', 'Consecutivo', 'Factura proveedor') para crear la nueva columna.")
        
        # 5. Extraer TRM de 'Observaciones' y sobrescribir 'Tasa de cambio'
        #if "Tasa de cambio" in df_procesado.columns and "Observaciones" in df_procesado.columns:
            #st.info("Actualizando 'Tasa de cambio' con los valores encontrados en 'Observaciones'...")

            #df_procesado['Observaciones'] = df_procesado['Observaciones'].astype(str)
            # Extrae el contenido de las llaves '{}'. El resultado será el texto o NaN si no hay llaves.
            #trm_extraida = df_procesado['Observaciones'].str.extract(r'\{(.*?)\}')[0]
            # Elimina las filas donde no se encontró nada (NaN), para quedarnos solo con los valores a actualizar.
            #trm_extraida.dropna(inplace=True)
            # Aseguramos que la columna 'Tasa de cambio' pueda recibir texto sin problemas.
            #df_procesado['Tasa de cambio'] = df_procesado['Tasa de cambio'].astype(object)
            # Actualiza la columna 'Tasa de cambio' SÓLO con los valores encontrados.
            # El método .update() alinea por índice y solo modifica donde hay coincidencia.
            #df_procesado['Tasa de cambio'].update(trm_extraida)
            
            #filas_actualizadas = len(trm_extraida)
            #st.success(f"Se actualizaron **{filas_actualizadas}** filas en 'Tasa de cambio'. Los valores existentes se respetaron donde no se encontró un valor entre {{}}.")
        #else:
            #st.warning("Advertencia: No se encontraron las columnas **'Tasa de cambio'** y/o **'Observaciones'**.")
        # 5. Extraer, LIMPIAR y sobrescribir 'Tasa de cambio' desde 'Observaciones' (LÓGICA CORREGIDA Y ENFOCADA)
        if "Tasa de cambio" in df_procesado.columns and "Observaciones" in df_procesado.columns:
            
            # Para evitar problemas, nos aseguramos de que la columna 'Tasa de cambio' sea numérica desde el principio.
            # Usamos la limpieza simple de comas que ya definimos.
            df_procesado['Tasa de cambio'] = convertir_a_numero_limpiando_comas(df_procesado['Tasa de cambio']).fillna(0)

            # 1. EXTRAER el valor de las observaciones como texto.
            trm_extraida = df_procesado['Observaciones'].astype(str).str.extract(r'\{(.*?)\}')[0]
            
            # Quitamos las filas donde no se encontró nada.
            trm_extraida.dropna(inplace=True)

            if not trm_extraida.empty:
                st.info("Valores de TRM encontrados en 'Observaciones'. Limpiando y actualizando...")

                # 2. LIMPIAR el texto extraído (quitamos comas de miles).
                # Ejemplo: "4,061.36" se convierte en "4061.36"
                trm_limpia = trm_extraida.str.replace(',', '', regex=False)

                # 3. CONVERTIR el texto limpio a un formato numérico.
                trm_numerica = pd.to_numeric(trm_limpia, errors='coerce')
                
                # Quitamos las filas donde la conversión a número pudo haber fallado.
                trm_numerica.dropna(inplace=True)

                # 4. ACTUALIZAR la columna 'Tasa de cambio' con los valores ya numéricos y limpios.
                # El método .update() alinea por índice y solo modifica donde encuentra correspondencia.
                df_procesado['Tasa de cambio'].update(trm_numerica)
                st.success(f"Se actualizaron **{len(trm_numerica)}** filas en 'Tasa de cambio' con valores numéricos limpios desde 'Observaciones'.")


        # 5.1. Calcular la nueva columna 'Valor Total ME' (VERSIÓN CORREGIDA FINAL)
        st.info("Calculando 'Valor Total ME'...")
        if 'Total' in df_procesado.columns and 'Tasa de cambio' in df_procesado.columns:
            
            # PASO CLAVE: Nos aseguramos de que 'Tasa de cambio' sea numérica OTRA VEZ,
            # justo antes de la división, para revertir el cambio a 'object' del paso anterior.
            tasa_numerica = pd.to_numeric(df_procesado['Tasa de cambio'], errors='coerce')
            
            # Reemplazamos 0 con NaN para evitar errores de división por cero.
            tasa_numerica.replace(0, np.nan, inplace=True)

            # Realizamos la división.
            df_procesado['Valor Total ME'] = df_procesado['Total'] / tasa_numerica
            
            # Rellenamos cualquier resultado inválido (NaN) con 0.
            df_procesado['Valor Total ME'].fillna(0, inplace=True)
            
            st.success("Se ha creado y calculado la columna **'Valor Total ME'**.")
        else:
            st.warning("No se pudo calcular 'Valor Total ME'.")

        # 6. Relacionar documentos FV-1 con DS-1 y FC-1
        st.info("Iniciando el proceso de relacionamiento de documentos...")
        
        # Separar el DataFrame en los dos grupos principales
        df_destino = df_procesado[df_procesado['Número comprobante'].isin(['FV-1', 'FV-2'])].copy()
        df_fuente = df_procesado[df_procesado['Número comprobante'].isin(['DS-1', 'FC-1'])].copy()

        if not df_fuente.empty:
            # Preparar el DataFrame fuente (DS-1, FC-1)
            df_fuente['NIT_relacion'] = df_fuente['Observaciones'].str.extract(r'\((.*?)\)')[0]
            
            df_destino['Identificación'] = df_destino['Identificación'].astype('Int64').astype(str)
            df_destino['Código'] = df_destino['Código'].astype(str)
            
            df_fuente['NIT_relacion'] = df_fuente['NIT_relacion'].astype(str)
            df_fuente['Código'] = df_fuente['Código'].astype(str)
            
            # Añadir prefijo a las columnas para evitar colisiones y dar claridad
            df_fuente = df_fuente.add_prefix('REL_')
            
            # Realizar la unión externa (outer join)
            df_final = pd.merge(
                df_destino,
                df_fuente,
                how='outer',
                left_on=['Identificación', 'Código'],
                right_on=['REL_NIT_relacion', 'REL_Código']
            )
            
            st.success("Relacionamiento completado. Los documentos sin pareja se han conservado.")
            df_procesado = df_final
        else:
            st.warning("No se encontraron documentos DS-1 o FC-1 para relacionar. El archivo final no tendrá columnas de relación.")
        
        # 7. Organizar y Limpiar Columnas Finales
        st.info("Organizando el formato final del archivo...")
        
        # A. Renombrar la columna "Tipo clasificación" a "Tipo Bien"
        # Verificamos si la columna existe antes de intentar renombrarla
        if "Tipo clasificación" in df_procesado.columns:
            df_procesado.rename(columns={"Tipo clasificación": "Tipo Bien"}, inplace=True)
            st.info("La columna **'Tipo clasificación'** ha sido renombrada a **'Tipo Bien'**.")
        
        if 'Tipo Bien' in df_procesado.columns:
            # Creamos un diccionario con los valores a reemplazar
            mapeo_valores = {
                'Servicio': 'S',
                'Producto': 'P'
            }
            df_procesado['Tipo Bien'].replace(mapeo_valores, inplace=True)
            st.info("Valores en 'Tipo Bien' actualizados: 'Servicio' a 'S' y 'Producto' a 'P'.")
        
        #Creación de la nueva columna "Vendedor"
        #if 'Vendedor' not in df_procesado.columns:
            #df_procesado['Vendedor'] = ''
        if 'Identificación Vendedor' in df_procesado.columns:
            df_procesado.drop(columns=['Identificación Vendedor'], inplace=True)
            
        #Creación de la nueva columna "Clasificación Producto"
        if 'Clasificación Producto' not in df_procesado.columns:
            df_procesado['Clasificación Producto'] = ''
            
        #Creación de la nueva columna "Línea"
        if 'Línea' not in df_procesado.columns:
            df_procesado['Línea'] = ''
            
        #Creación de la nueva columna "Descripción Línea"
        if 'Descripción Línea' not in df_procesado.columns:
            df_procesado['Descripción Línea'] = ''
            
        #Creación de la nueva columna "Sublínea"
        if 'Sublínea' not in df_procesado.columns:
            df_procesado['Sublínea'] = ''
            
        #Creación de la nueva columna "Descripción Sublínea"
        if 'Descripción Sublínea' not in df_procesado.columns:
            df_procesado['Descripción Sublínea'] = ''
            
        
        #Se define el orden y la selección final de las columnas
        columnas_finales = [
            # Columnas del lado izquierdo (FV)
            'Tipo Bien', 'Clasificación Producto', 'Línea', 'Descripción Línea', 'Sublínea', 'Descripción Sublínea', 'Código', 'Nombre', 'Número comprobante', 'Numero comprobante',
            'Fecha elaboración', 'Identificación', 'Nombre tercero', 'Vendedor', 'Cantidad',
            'Valor unitario', 'Total', 'Tasa de cambio', 'Valor Total ME', 'Observaciones',
            
            # Columnas del lado derecho (REL_)
            'REL_Número comprobante', 'REL_Consecutivo',
            'REL_Factura proveedor', 'REL_Identificación', 'REL_Nombre tercero', 'REL_Cantidad',
            'REL_Valor unitario',  'REL_Tasa de cambio', 'REL_Total', 'REL_Valor Total ME'
        ]
        
        # Filtrar la lista para incluir solo las columnas que realmente existen en el DataFrame
        # Esto hace el código más robusto si alguna columna faltara
        columnas_existentes_ordenadas = [col for col in columnas_finales if col in df_procesado.columns]

        # Reordenar y eliminar las columnas no deseadas de una sola vez
        df_procesado = df_procesado[columnas_existentes_ordenadas]

        st.success("Columnas reorganizadas y limpiadas con éxito.")
 
        st.success("¡Procesamiento completado con éxito!")
        
        # --- SEGUNDA VERIFICACIÓN DE DUPLICADOS (Interna del DataFrame) ---
        st.info("Ejecutando segunda verificación de duplicados internos...")
        
        # Definimos las columnas que identifican un registro único según tu regla
        columnas_unicas = [
            'REL_Factura proveedor', 
            'Código', 
            'REL_Nombre tercero', 
            'Nombre tercero', 
            'Identificación', 
            'Fecha elaboración'
        ]
        
        # Verificamos cuáles de estas columnas existen realmente en el DF para evitar errores
        cols_presentes = [c for c in columnas_unicas if c in df_procesado.columns]
        
        filas_antes_segunda_limpieza = len(df_procesado)
        
        # Eliminamos duplicados dejando solo la primera aparición
        df_procesado = df_procesado.drop_duplicates(subset=cols_presentes, keep='first')
        
        filas_despues_segunda_limpieza = len(df_procesado)
        duplicados_internos = filas_antes_segunda_limpieza - filas_despues_segunda_limpieza
        
        if duplicados_internos > 0:
            st.warning(f"✅ Se eliminaron {duplicados_internos} registros duplicados encontrados dentro del mismo proceso.")
        else:
            st.success("✅ No se encontraron duplicados internos en esta ejecución.")
        # -----------------------------------------------------------------
        
        return df_procesado

    except Exception as e:
        st.error(f"Se produjo un error durante el procesamiento: {e}")
        return None

# --- Interfaz de Usuario de Streamlit ---
st.set_page_config(page_title="Procesador de Excel Automático", layout="centered")

st.title("📊 Procesador de Archivos Excel")
st.markdown("---")

uploaded_file = st.file_uploader(
    "Sube tu archivo Excel (.xlsx)",
    type=["xlsx"],
    help="Arrastra y suelta tu archivo Excel aquí o haz clic para buscar."
)

#st.markdown("---")
#st.header("🔧 Herramientas de Debug para SharePoint")

# Crear variables de prueba para conexión SharePoint
#if st.button("🔗 Probar Conexión SharePoint (Solo Debug)"):
    #with st.spinner("Conectando..."):
        #status_placeholder = st.empty()
        #token = get_access_token(status_placeholder)
        
        #if token:
            #site_id = get_sharepoint_site_id(token)
            #if site_id:
                #headers = {'Authorization': f'Bearer {token}'}
                #st.session_state.debug_headers = headers
                #st.session_state.debug_site_id = site_id
                #st.success("✅ Conexión establecida para debug")

# Solo mostrar herramientas de debug si hay conexión
if hasattr(st.session_state, 'debug_headers') and hasattr(st.session_state, 'debug_site_id'):
    
    with st.expander("🧪 Debug de Archivos SharePoint", expanded=False):
        
        # Debug para la carpeta mensual
        #st.subheader("📅 Debug de Carpeta Mensual")
        if st.button("Listar archivos en carpeta mensual"):
            listar_archivos_en_carpeta(st.session_state.debug_headers, st.session_state.debug_site_id, "Ventas con ciudad 2025")
        
        # Debug para archivo específico
        #st.subheader("🔍 Debug de Archivo Específico")
        archivo_debug = st.text_input("Ruta completa del archivo a verificar:", 
                                     "Ventas con ciudad 2025/Ventas Septiembre 2025.xlsx")
        
        col1, col2 = st.columns(2)
        with col1:
            if st.button("Verificar archivo"):
                if archivo_debug:
                    verificar_archivo_existe_sharepoint(st.session_state.debug_headers, st.session_state.debug_site_id, archivo_debug)
        
        with col2:
            if st.button("Intentar descargar"):
                if archivo_debug:
                    contenido = obtener_contenido_archivo_sharepoint(st.session_state.debug_headers, st.session_state.debug_site_id, archivo_debug)
                    if contenido:
                        st.success(f"Archivo descargado exitosamente ({len(contenido):,} bytes)")

st.markdown("---")

df_result = None


if uploaded_file is not None:
    st.success(f"Archivo **'{uploaded_file.name}'** cargado correctamente.")
    
    if st.button("Iniciar Procesamiento"):
        # Crear el placeholder una sola vez
        status_placeholder = st.empty()
        with st.spinner("Procesando tu archivo... Esto puede tardar unos minutos, especialmente al consultar la TRM..."):
            df_result = procesar_excel_para_streamlit(uploaded_file, status_placeholder)
        
        if df_result is not None:
            st.subheader("Vista previa del archivo procesado:")
            st.dataframe(df_result.head())

            output = io.BytesIO()
            # 2. Conectarse a SharePoint
            token = get_access_token(status_placeholder)
            if token:
                
                site_id = get_sharepoint_site_id(token) # Esta función es rápida, no necesita placeholder

                if site_id:
                    # Una vez que tenemos el site_id, AHORA creamos los headers para las siguientes funciones
                    headers = {'Authorization': f'Bearer {token}'}
                    # 3. Encontrar el archivo del mes
                    ruta_archivo_mensual = encontrar_archivo_del_mes(headers, site_id, RUTA_CARPETA_VENTAS_MENSUALES, status_placeholder)
                    ruta_fija_trm = "01 Archivos Area Administrativa/TRM4.xlsx"
                    exito_trm = actualizar_archivo_trm(headers, site_id, ruta_fija_trm, df_result, status_placeholder)
                    #st.info("Archivo TRM actualizado con Éxito")
                    if ruta_archivo_mensual:
                        # 4. Agregar los datos
                        #agregar_datos_a_excel_sharepoint(headers, site_id, ruta_archivo_mensual, df_result, status_placeholder)
                        exito = agregar_datos_a_excel_sharepoint(headers, site_id, ruta_archivo_mensual, df_result, status_placeholder)
                        if exito:
                            st.balloons()
            
            with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
                df_result.to_excel(writer, index=False, sheet_name='Procesado')
            processed_data = output.getvalue()

            st.download_button(
                label="Descargar Archivo Procesado",
                data=processed_data,
                file_name=f"procesado_{uploaded_file.name}",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            st.info("Tu archivo ha sido procesado y está listo para descargar.")
else:
    st.info("Por favor, sube un archivo Excel para comenzar.")